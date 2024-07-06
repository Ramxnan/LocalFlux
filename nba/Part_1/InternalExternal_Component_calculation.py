from openpyxl import *
from openpyxl.utils import get_column_letter
from .utils import cellstyle, cellstyle_range
from openpyxl.styles import Protection


def Component_calculation(data,Component_Details,aw,component_type):
    ''' Function to create the component calculation table

    Args:
    data (dict): Dictionary containing the data
    Component_Details (dict): Dictionary containing the component details
    aw (openpyxl.worksheet.worksheet.Worksheet): Worksheet object
    component_type (str): Component type

    Returns:
    openpyxl.worksheet.worksheet.Worksheet: Worksheet object
    '''

    for row in aw.iter_rows():
        for cell in row:
            cell.protection = Protection(locked=True)
    aw.protection.sheet = True

    #get the value from other sheets and put it in this sheet as a formula dynamically by referencing the cell
    start_column=1
    component_num=2
    components_len=0
    for component_name in Component_Details.keys():
        if component_name[-1]==component_type:
            aw[f'{get_column_letter(start_column)}1']=component_name
            aw.merge_cells(start_row=1, start_column=start_column, end_row=1, end_column=start_column+data['Number_of_COs']-1)
            cellstyle_range(aw[f'{get_column_letter(start_column)}1:{get_column_letter(start_column+data["Number_of_COs"]-1)}1'], bold=True, alignment=True, border=True, fill="4f81bd", size=12)
            

            #reference a cell from another sheet
            for nco in range(1,data['Number_of_COs']+1):
                aw[f'{get_column_letter(start_column+nco-1)}2']=f"CO{nco}"
                cellstyle(aw[f'{get_column_letter(start_column+nco-1)}2'],fill="4f81bd")

                aw[f'{get_column_letter(start_column+nco-1)}3']=f"='{component_name}'!{get_column_letter(2+Component_Details[component_name]+1+nco)}3"

                aw[f'{get_column_letter(start_column+nco-1)}4']=f"='{component_name}'!{get_column_letter(2+Component_Details[component_name]+1+nco)}4"

            cellstyle_range(aw[f'{get_column_letter(start_column)}2:{get_column_letter(start_column+data["Number_of_COs"]-1)}4'], alignment=True, border=True, bold=True)
                
            
            for nco in range(1,data['Number_of_COs']+1):
                aw[f'{get_column_letter(start_column+nco-1)}6']=f"CO{nco}"
                cellstyle(aw[f'{get_column_letter(start_column+nco-1)}6'],fill="4f81bd")

                for nstudents in range(1,data['Number_of_Students']+1):
                    aw[f'{get_column_letter(start_column+nco-1)}{6+nstudents}']=f"='{component_name}'!{get_column_letter(2+Component_Details[component_name]+1+nco)}{10+nstudents}"
            
            cellstyle_range(aw[f'{get_column_letter(start_column)}6:{get_column_letter(start_column+data["Number_of_COs"]-1)}{6+data["Number_of_Students"]}'], alignment=True, border=True, bold=True)
                

            start_column+=data['Number_of_COs']+1 
            component_num+=1
            components_len+=1

    #make a column of rows black to demarcate the components
    #set width of columns
    aw.column_dimensions[f'{get_column_letter(start_column)}'].width = 2.5
    cellstyle_range(aw[f'{get_column_letter(start_column)}1:{get_column_letter(start_column)}{data["Number_of_Students"]+11}'], fill="000000")


    start_column+=2
    #set header to Combined Components
    aw[f'{get_column_letter(start_column)}1']="Combined Components table"
    aw.merge_cells(start_row=1, start_column=start_column, end_row=1, end_column=start_column+data['Number_of_COs']-1)
    cellstyle_range(aw[f'{get_column_letter(start_column)}1:{get_column_letter(start_column+data["Number_of_COs"]-1)}1'], bold=True, alignment=True, border=True, fill="000000",font_color="FFFFFF", size=12)

    for nco in range(1,data['Number_of_COs']+1):
        aw[f'{get_column_letter(start_column+nco-1)}2']=f"CO{nco}"
        cellstyle(aw[f'{get_column_letter(start_column+nco-1)}2'],bold=True,fill="000000",font_color="FFFFFF",alignment=True)
        
        corr_co_column=nco
        formula="=SUM("
        for _ in range(1,components_len+1):
            formula+=f"{get_column_letter(corr_co_column)}3,"
            corr_co_column+=data['Number_of_COs']+1
        formula=formula[:-1]
        formula+=")"
        aw[f'{get_column_letter(start_column+nco-1)}3']=formula

        corr_co_column=nco
        formula="=SUM("
        for _ in range(1,components_len+1):
            formula+=f"{get_column_letter(corr_co_column)}4,"
            corr_co_column+=data['Number_of_COs']+1
        formula=formula[:-1]
        formula+=")"
        aw[f'{get_column_letter(start_column+nco-1)}4']=formula


    cellstyle_range(aw[f'{get_column_letter(start_column)}3:{get_column_letter(start_column+data["Number_of_COs"]-1)}4'], alignment=True, border=True, bold=True)


    #Combined Marks
    for nco in range(1,data['Number_of_COs']+1):
        aw[f'{get_column_letter(start_column+nco-1)}6']=f"CO{nco}"
        cellstyle(aw[f'{get_column_letter(start_column+nco-1)}6'],bold=True,fill="000000",font_color="FFFFFF",alignment=True)

        for nstudents in range(1,data['Number_of_Students']+1):
            corr_co_column=nco
            formula="=SUM("
            for _ in range(1,components_len+1):
                formula+=f"{get_column_letter(corr_co_column)}{6+nstudents},"
                corr_co_column+=data['Number_of_COs']+1
            formula=formula[:-1]
            formula+=")"

            aw[f'{get_column_letter(start_column+nco-1)}{6+nstudents}']=formula
            
    cellstyle_range(aw[f'{get_column_letter(start_column)}7:{get_column_letter(start_column+data["Number_of_COs"]-1)}{6+data["Number_of_Students"]}'], alignment=True, border=True, bold=True)

    #Total final calculation
    aw.column_dimensions[f'{get_column_letter(start_column-1)}'].width = 14.3
    aw[f'{get_column_letter(start_column-1)}{data["Number_of_Students"]+8}']="CO"
    aw[f'{get_column_letter(start_column-1)}{data["Number_of_Students"]+9}']="CO%"
    aw[f'{get_column_letter(start_column-1)}{data["Number_of_Students"]+10}']="Total students"
    if component_type=="I":
            aw[f'{get_column_letter(start_column-1)}{data["Number_of_Students"]+11}']="I-attainment %"
    else:
        aw[f'{get_column_letter(start_column-1)}{data["Number_of_Students"]+11}']="E-attainment %"
    
    cellstyle_range(aw[f'{get_column_letter(start_column-1)}{data["Number_of_Students"]+8}:{get_column_letter(start_column-1)}{data["Number_of_Students"]+11}'], alignment=True, border=True, bold=True, fill="000000",font_color="FFFFFF")

    for nco in range(1,data['Number_of_COs']+1):
        aw[f'{get_column_letter(start_column-1+nco)}{data["Number_of_Students"]+8}']=f"CO{nco}"
        cellstyle(aw[f'{get_column_letter(start_column-1+nco)}{data["Number_of_Students"]+8}'],bold=True,fill="000000",font_color="FFFFFF",alignment=True)

        range_start = f"{get_column_letter(start_column-1+nco)}7"
        range_end = f"{get_column_letter(start_column-1+nco)}{6+data['Number_of_Students']}"
        range_string = f"{range_start}:{range_end}"

        # Building the COUNTIF formula with an embedded IF condition
        criteria_cell = f"{get_column_letter(start_column-1+nco)}4"
        criteria = f'">=" & {criteria_cell}'

        # The IF condition checks if the sum of the range is greater than zero (indicating non-zero values are present)
        formula = f'=IF(SUM({range_string}) > 0, COUNTIF({range_string}, {criteria}), "")'  # Empty string or "N/A" as placeholder

        # Write the formula to the cell
        aw[f'{get_column_letter(start_column-1+nco)}{data["Number_of_Students"]+9}'] = formula

        aw[f'{get_column_letter(start_column-1+nco)}{data["Number_of_Students"]+10}']=data['Number_of_Students']
        
        num_students = data['Number_of_Students']
        cell_position_1 = f"{get_column_letter(start_column-1+nco)}{num_students+9}"
        cell_position_2 = f"{get_column_letter(start_column-1+nco)}{num_students+10}"
        target_cell_position = f"{get_column_letter(start_column-1+nco)}{num_students+11}"
        formula = f'=IF(SUM({range_string}) > 0, {cell_position_1}/{cell_position_2}*100, "0")'
        aw[target_cell_position] = formula

    cellstyle_range(aw[f'{get_column_letter(start_column)}{data["Number_of_Students"]+9}:{get_column_letter(start_column+data["Number_of_COs"]-1)}{data["Number_of_Students"]+11}'], alignment=True, border=True, bold=True, alternate=['ffffff','d9d9d9'])

    return aw

