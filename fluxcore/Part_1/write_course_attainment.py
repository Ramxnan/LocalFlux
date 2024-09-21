import pandas as pd
from openpyxl.styles import Alignment, PatternFill, Protection
from openpyxl.utils import get_column_letter
from .Input_Details import input_detail, indirect_co_assessment, CO_PO_Table
from .utils import cellstyle, cellstyle_range, adjust_width

def write_course_attainment(data,Component_Details,config,aw):
    # for row in aw.iter_rows():
    #     for cell in row:
    #         cell.protection = Protection(locked=True)
    # aw.protection.sheet = True

    #============================================================================================================
    aw=input_detail(data,Component_Details,aw,copy=True)
    adjust_width(aw)

    #============================================================================================================
    #============================================================================================================
    aw=CO_PO_Table(data,config,aw,copy=True)
    

    #============================================================================================================
    #============================================================================================================
    aw=indirect_co_assessment(data,aw,copy=True,ca_page=True)
                
    #============================================================================================================
    #============================================================================================================
    start_range_1 = f'G{data["Number_of_COs"]+7}'
    start_range_2 = f'G{data["Number_of_COs"]+8}'
    start_range_3 = f'G{data["Number_of_COs"]+9}'
    end_range_1 = f'H{data["Number_of_COs"]+7}'
    end_range_2 = f'H{data["Number_of_COs"]+8}'
    end_range_3 = f'H{data["Number_of_COs"]+9}'
    map_1 = f'I{data["Number_of_COs"]+7}'
    map_2 = f'I{data["Number_of_COs"]+8}'
    map_3 = f'I{data["Number_of_COs"]+9}'


    #merge cells col 4 to 16 and rows 8 to 10
    start_col_ca = 4
    start_row_ca = data["Number_of_COs"]+8+data["Number_of_COs"]+1

    aw.merge_cells(start_row=start_row_ca, start_column=start_col_ca, end_row=start_row_ca+2, end_column=start_col_ca+12)
    aw[f'{get_column_letter(start_col_ca)}{start_row_ca}'] = 'Course Attainment'
    cellstyle_range(aw[f'{get_column_letter(start_col_ca)}{start_row_ca}:{get_column_letter(start_col_ca+12)}{start_row_ca+2}'],border=True, fill="ffe74e", bold=True, alignment=True)

    start_row_ca+=3
    aw.merge_cells(start_row=start_row_ca, start_column=start_col_ca, end_row=start_row_ca+3, end_column=start_col_ca)
    aw[f'{get_column_letter(start_col_ca)}{start_row_ca}'] = 'Course Outcome'

    aw.merge_cells(start_row=start_row_ca, start_column=start_col_ca+1, end_row=start_row_ca, end_column=start_col_ca+2)
    aw[f'{get_column_letter(start_col_ca+1)}{start_row_ca}'] = 'Mapping with Program'

    aw.merge_cells(start_row=start_row_ca+1, start_column=start_col_ca+1, end_row=start_row_ca+3, end_column=start_col_ca+1)
    aw[f'{get_column_letter(start_col_ca+1)}{start_row_ca+1}'] = 'POs & PSOs'

    aw[f'{get_column_letter(start_col_ca+2)}{start_row_ca+1}'] = 'Level of Mapping'

    aw.merge_cells(start_row=start_row_ca+2, start_column=start_col_ca+2, end_row=start_row_ca+3, end_column=start_col_ca+2)
    aw[f'{get_column_letter(start_col_ca+2)}{start_row_ca+2}'] = 'Affinity'

    aw.merge_cells(start_row=start_row_ca, start_column=start_col_ca+3, end_row=start_row_ca, end_column=start_col_ca+12)
    aw[f'{get_column_letter(start_col_ca+3)}{start_row_ca}'] = 'Attainment % in'

    aw.merge_cells(start_row=start_row_ca+1, start_column=start_col_ca+3, end_row=start_row_ca+1, end_column=start_col_ca+8)
    aw[f'{get_column_letter(start_col_ca+3)}{start_row_ca+1}'] = 'Direct'

    aw.merge_cells(start_row=start_row_ca+2, start_column=start_col_ca+3, end_row=start_row_ca+2, end_column=start_col_ca+4)
    aw[f'{get_column_letter(start_col_ca+3)}{start_row_ca+2}'] = 'University(SEE)'

    aw[f'{get_column_letter(start_col_ca+3)}{start_row_ca+3}'] = 'Attainment'

    # aw[f'{get_column_letter(start_col_ca+4)}{start_row_ca+3}'] = 'Level Of Attainment (0-40 --> 1, 40-60 ---> 2, 60-100---> 3)'
    target_cell=aw[f'{get_column_letter(start_col_ca+4)}{start_row_ca+3}']
    target_cell.value=(
        f'=CONCATENATE("Level Of Attainment (", {start_range_1}, "-", {end_range_1}, " --> ", {map_1}, ", ", '
        f'{start_range_2}, "-", {end_range_2}, " --> ", {map_2}, ", ", '
        f'{start_range_3}, "-", {end_range_3}, " --> ", {map_3}, ")")'
        )
    


    aw.merge_cells(start_row=start_row_ca+2, start_column=start_col_ca+5, end_row=start_row_ca+2, end_column=start_col_ca+6)
    aw[f'{get_column_letter(start_col_ca+5)}{start_row_ca+2}'] = 'Internal(CIE)'

    aw[f'{get_column_letter(start_col_ca+5)}{start_row_ca+3}'] = 'Attainment'

    # aw[f'{get_column_letter(start_col_ca+6)}{start_row_ca+3}'] = 'Level Of Attainment (0-40 --> 1, 40-60 ---> 2, 60-100---> 3)'
    target_cell=aw[f'{get_column_letter(start_col_ca+6)}{start_row_ca+3}']
    target_cell.value=(
        f'=CONCATENATE("Level Of Attainment (", {start_range_1}, "-", {end_range_1}, " --> ", {map_1}, ", ", '
        f'{start_range_2}, "-", {end_range_2}, " --> ", {map_2}, ", ", '
        f'{start_range_3}, "-", {end_range_3}, " --> ", {map_3}, ")")'
        )
    

    aw.merge_cells(start_row=start_row_ca+2, start_column=start_col_ca+7, end_row=start_row_ca+2, end_column=start_col_ca+8)
    aw[f'{get_column_letter(start_col_ca+7)}{start_row_ca+2}'] = '="Weighted Level of Attainment (" & B16 & " SEE + " & B15 & " CIE)"'
    aw.row_dimensions[start_row_ca+2].height = 52

    aw[f'{get_column_letter(start_col_ca+7)}{start_row_ca+3}'] = 'Attainment'

    # aw[f'{get_column_letter(start_col_ca+8)}{start_row_ca+3}'] = 'Level Of Attainment (0-40 --> 1, 40-60 ---> 2, 60-100---> 3)'
    target_cell=aw[f'{get_column_letter(start_col_ca+8)}{start_row_ca+3}']
    target_cell.value=(
        f'=CONCATENATE("Level Of Attainment (", {start_range_1}, "-", {end_range_1}, " --> ", {map_1}, ", ", '
        f'{start_range_2}, "-", {end_range_2}, " --> ", {map_2}, ", ", '
        f'{start_range_3}, "-", {end_range_3}, " --> ", {map_3}, ")")'
        )

    aw.merge_cells(start_row=start_row_ca+1, start_column=start_col_ca+9, end_row=start_row_ca+1, end_column=start_col_ca+10)
    aw[f'{get_column_letter(start_col_ca+9)}{start_row_ca+1}'] = 'Indirect'

    aw.merge_cells(start_row=start_row_ca+2, start_column=start_col_ca+9, end_row=start_row_ca+3, end_column=start_col_ca+9)
    aw[f'{get_column_letter(start_col_ca+9)}{start_row_ca+2}'] = 'Attainment'

    aw.merge_cells(start_row=start_row_ca+2, start_column=start_col_ca+10, end_row=start_row_ca+3, end_column=start_col_ca+10)
    # aw[f'{get_column_letter(start_col_ca+10)}{start_row_ca+2}'] = 'Level Of Attainment (0-40 --> 1, 40-60 ---> 2, 60-100---> 3)'
    target_cell=aw[f'{get_column_letter(start_col_ca+10)}{start_row_ca+2}']
    target_cell.value=(
        f'=CONCATENATE("Level Of Attainment (", {start_range_1}, "-", {end_range_1}, " --> ", {map_1}, ", ", '
        f'{start_range_2}, "-", {end_range_2}, " --> ", {map_2}, ", ", '
        f'{start_range_3}, "-", {end_range_3}, " --> ", {map_3}, ")")'
        )

    aw.merge_cells(start_row=start_row_ca+1, start_column=start_col_ca+11, end_row=start_row_ca+2, end_column=start_col_ca+12)
    aw[f'{get_column_letter(start_col_ca+11)}{start_row_ca+1}'] = 'Final Weighted CO Attainment (80% Direct + 20% Indirect)'

    aw[f'{get_column_letter(start_col_ca+11)}{start_row_ca+3}'] = 'Attainment'

    # aw[f'{get_column_letter(start_col_ca+12)}{start_row_ca+3}'] = 'Level Of Attainment (0-40 --> 1, 40-60 ---> 2, 60-100---> 3)'
    target_cell=aw[f'{get_column_letter(start_col_ca+12)}{start_row_ca+3}']
    target_cell.value=(
        f'=CONCATENATE("Level Of Attainment (", {start_range_1}, "-", {end_range_1}, " --> ", {map_1}, ", ", '
        f'{start_range_2}, "-", {end_range_2}, " --> ", {map_2}, ", ", '
        f'{start_range_3}, "-", {end_range_3}, " --> ", {map_3}, ")")'
        )

    aw.column_dimensions[get_column_letter(start_col_ca-1)].width = 17.22
    aw.column_dimensions[get_column_letter(start_col_ca)].width = 17.22
    aw.column_dimensions[get_column_letter(start_col_ca+1)].width = 9.33
    aw.column_dimensions[get_column_letter(start_col_ca+2)].width = 15.56
    aw.column_dimensions[get_column_letter(start_col_ca+3)].width = 13
    aw.column_dimensions[get_column_letter(start_col_ca+4)].width = 12
    aw.column_dimensions[get_column_letter(start_col_ca+5)].width = 13
    aw.column_dimensions[get_column_letter(start_col_ca+6)].width = 12
    aw.column_dimensions[get_column_letter(start_col_ca+7)].width = 13
    aw.column_dimensions[get_column_letter(start_col_ca+8)].width = 12
    aw.column_dimensions[get_column_letter(start_col_ca+9)].width = 13
    aw.column_dimensions[get_column_letter(start_col_ca+10)].width = 12
    aw.column_dimensions[get_column_letter(start_col_ca+11)].width = 13
    aw.column_dimensions[get_column_letter(start_col_ca+12)].width = 12
    
    cellstyle_range(aw[f'{get_column_letter(start_col_ca)}{start_row_ca}:{get_column_letter(start_col_ca+12)}{start_row_ca+3}'],border=True, fill="b8cce4", bold=True, alignment=True, wrap_text=True)
            
    cellstyle(aw[f'{get_column_letter(start_col_ca+2)}{start_row_ca+2}'], fill="c4d79b")
    cellstyle(aw[f'{get_column_letter(start_col_ca+7)}{start_row_ca+2}'], fill="c4d79b")
    cellstyle(aw[f'{get_column_letter(start_col_ca+7)}{start_row_ca+3}'], fill="c4d79b")
    cellstyle(aw[f'{get_column_letter(start_col_ca+8)}{start_row_ca+3}'], fill="c4d79b")
    cellstyle(aw[f'{get_column_letter(start_col_ca+11)}{start_row_ca+1}'], fill="c4d79b")
    cellstyle(aw[f'{get_column_letter(start_col_ca+11)}{start_row_ca+3}'], fill="c4d79b")
    cellstyle(aw[f'{get_column_letter(start_col_ca+12)}{start_row_ca+3}'], fill="c4d79b")


    #================================================================================================================================================================

    start=start_row_ca+3
    # interval=16
    interval=config["PO"]+config["PSO"] -1
    rowindex=1
    for i in range(1, (data["Number_of_COs"]+1)):
        start+=1
        aw.merge_cells(start_row=start, start_column=start_col_ca, end_row=start+interval, end_column=start_col_ca)
        aw.cell(row=start, column=start_col_ca).value = "CO"+str(i)
        cellstyle(aw.cell(row=start, column=start_col_ca), bold=True)
        aw.cell(row=start, column=start_col_ca).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if i%2==0:
            aw.cell(row=start, column=start_col_ca).fill = PatternFill(start_color='c4d79b', end_color='c4d79b', fill_type='solid')
        else:
            aw.cell(row=start, column=start_col_ca).fill = PatternFill(start_color='b8cce4', end_color='b8cce4', fill_type='solid')


        index=1
        for j in range(start, start+interval+1):
            #print out COPO mapping
            aw.cell(row=j, column=start_col_ca+1).value =f'={get_column_letter(index+4)}2'
            
            
            aw.cell(row=j, column=start_col_ca+2).value = f'={get_column_letter(index+4)}{i+2}'
            
            
            if index%2==0:
                cellstyle_range(aw[f'{get_column_letter(start_col_ca+1)}{j}:{get_column_letter(start_col_ca+2)}{j}'], fill="c4d79b")
            else:
                cellstyle_range(aw[f'{get_column_letter(start_col_ca+1)}{j}:{get_column_letter(start_col_ca+2)}{j}'], fill="b8cce4")
            index+=1

        for k in range(start_col_ca+3, start_col_ca+13):
            aw.merge_cells(start_row=start, start_column=k, end_row=start+interval, end_column=k)
            if k%2==0:
                cellstyle(aw.cell(row=start, column=k), fill="b8cce4")
            else:
                cellstyle(aw.cell(row=start, column=k), fill="dce6f1")
    
        internal_components_number = len([key for key in Component_Details.keys() if key.endswith("I")])
        external_components_number = len([key for key in Component_Details.keys() if key.endswith("E")])
        col=(data["Number_of_COs"]*external_components_number) + (1*external_components_number) + 2 + rowindex
        row=(6 + data["Number_of_Students"] + 5)
        aw.cell(row=start, column=start_col_ca+3).value=f'={data["Section"]}_External_Components!{get_column_letter(col)}{row}'




        # aw.cell(row=start, column=start_col_ca+4).value=f'=IF(AND({get_column_letter(start_col_ca+3)}{start}>0,{get_column_letter(start_col_ca+3)}{start}<40),1,IF(AND({get_column_letter(start_col_ca+3)}{start}>=40,{get_column_letter(start_col_ca+3)}{start}<60),2,IF(AND({get_column_letter(start_col_ca+3)}{start}>=60,{get_column_letter(start_col_ca+3)}{start}<=100),3,"0")))'
        # aw.cell(row=start, column=start_col_ca+4).value=f'=IF(AND({get_column_letter(start_col_ca+3)}{start}>{start_range_1},{get_column_letter(start_col_ca+3)}{start}<{end_range_1}),{map_1},IF(AND({get_column_letter(start_col_ca+3)}{start}>{start_range_2},{get_column_letter(start_col_ca+3)}{start}<{end_range_2}),{map_2},IF(AND({get_column_letter(start_col_ca+3)}{start}>{start_range_3},{get_column_letter(start_col_ca+3)}{start}<{end_range_3}),{map_3},"0")))'
        target_cell=aw.cell(row=start, column=start_col_ca+4)
        target_cell.value = (
            f'=IF(AND({get_column_letter(start_col_ca+3)}{start}>={start_range_1}, '
            f'{get_column_letter(start_col_ca+3)}{start}<={end_range_1}), {map_1}, '
            f'IF(AND({get_column_letter(start_col_ca+3)}{start}>{end_range_1}, '
            f'{get_column_letter(start_col_ca+3)}{start}<={end_range_2}), {map_2}, '
            f'IF(AND({get_column_letter(start_col_ca+3)}{start}>{end_range_2}, '
            f'{get_column_letter(start_col_ca+3)}{start}<={end_range_3}), {map_3}, "0")))'
        )



        col=(data["Number_of_COs"]*internal_components_number) + (1*internal_components_number) + 2 + rowindex
        row=(6 + data["Number_of_Students"] + 5)
        aw.cell(row=start, column=start_col_ca+5).value=f'={data["Section"]}_Internal_Components!{get_column_letter(col)}{row}'

        # aw.cell(row=start, column=start_col_ca+6).value=f'=IF(AND({get_column_letter(start_col_ca+5)}{start}>0,{get_column_letter(start_col_ca+5)}{start}<40),1,IF(AND({get_column_letter(start_col_ca+5)}{start}>=40,{get_column_letter(start_col_ca+5)}{start}<60),2,IF(AND({get_column_letter(start_col_ca+5)}{start}>=60,{get_column_letter(start_col_ca+5)}{start}<=100),3,"0")))'
        # aw.cell(row=start, column=start_col_ca+6).value=f'=IF(AND({get_column_letter(start_col_ca+3)}{start}>{start_range_1},{get_column_letter(start_col_ca+3)}{start}<{end_range_1}),{map_1},IF(AND({get_column_letter(start_col_ca+3)}{start}>{start_range_2},{get_column_letter(start_col_ca+3)}{start}<{end_range_2}),{map_2},IF(AND({get_column_letter(start_col_ca+3)}{start}>{start_range_3},{get_column_letter(start_col_ca+3)}{start}<{end_range_3}),{map_3},"0")))'
        target_cell=aw.cell(row=start, column=start_col_ca+6)
        target_cell.value=(
            f'=IF(AND({get_column_letter(start_col_ca+5)}{start}>={start_range_1},'
            f'{get_column_letter(start_col_ca+5)}{start}<={end_range_1}),{map_1},'
            f'IF(AND({get_column_letter(start_col_ca+5)}{start}>{end_range_1},'
            f'{get_column_letter(start_col_ca+5)}{start}<={end_range_2}),{map_2},'
            f'IF(AND({get_column_letter(start_col_ca+5)}{start}>{end_range_2},'
            f'{get_column_letter(start_col_ca+5)}{start}<={end_range_3}),{map_3},"0")))'
        )


        SEE_attainment = f'{get_column_letter(start_col_ca+3)}{start}'
        CIE_attainment = f'{get_column_letter(start_col_ca+5)}{start}'
        if_formula = (
            f'IF(OR(VALUE({SEE_attainment})=0, VALUE({CIE_attainment})=0), '
            f'IF(VALUE({SEE_attainment})=0, VALUE({CIE_attainment}), VALUE({SEE_attainment})), '
            f'VALUE({SEE_attainment})*(B16/100) + VALUE({CIE_attainment})*(B15/100))'
        )
        calculation = if_formula


        formula = f'={calculation}'
        aw.cell(row=start, column=start_col_ca+7).value = formula

        # aw.cell(row=start, column=start_col_ca+8).value=f'=IF(AND({get_column_letter(start_col_ca+7)}{start}>0,{get_column_letter(start_col_ca+7)}{start}<40),1,IF(AND({get_column_letter(start_col_ca+7)}{start}>=40,{get_column_letter(start_col_ca+7)}{start}<60),2,IF(AND({get_column_letter(start_col_ca+7)}{start}>=60,{get_column_letter(start_col_ca+7)}{start}<=100),3,"0")))'
        # aw.cell(row=start, column=start_col_ca+8).value=f'=IF(AND({get_column_letter(start_col_ca+3)}{start}>{start_range_1},{get_column_letter(start_col_ca+3)}{start}<{end_range_1}),{map_1},IF(AND({get_column_letter(start_col_ca+3)}{start}>{start_range_2},{get_column_letter(start_col_ca+3)}{start}<{end_range_2}),{map_2},IF(AND({get_column_letter(start_col_ca+3)}{start}>{start_range_3},{get_column_letter(start_col_ca+3)}{start}<{end_range_3}),{map_3},"0")))'
        target_cell=aw.cell(row=start, column=start_col_ca+8)
        target_cell.value=(
            f'=IF(AND({get_column_letter(start_col_ca+7)}{start}>={start_range_1},'
            f'{get_column_letter(start_col_ca+7)}{start}<={end_range_1}),{map_1},'
            f'IF(AND({get_column_letter(start_col_ca+7)}{start}>{end_range_1},'
            f'{get_column_letter(start_col_ca+7)}{start}<={end_range_2}),{map_2},'
            f'IF(AND({get_column_letter(start_col_ca+7)}{start}>{end_range_2},'
            f'{get_column_letter(start_col_ca+7)}{start}<={end_range_3}),{map_3},"0")))'
        )


        aw.cell(row=start, column=start_col_ca+9).value=f'=E{2+data["Number_of_COs"]+4+rowindex}'

        # aw.cell(row=start, column=start_col_ca+10).value=f'=IF(AND({get_column_letter(start_col_ca+9)}{start}>0,{get_column_letter(start_col_ca+9)}{start}<40),1,IF(AND({get_column_letter(start_col_ca+9)}{start}>=40,{get_column_letter(start_col_ca+9)}{start}<60),2,IF(AND({get_column_letter(start_col_ca+9)}{start}>=60,{get_column_letter(start_col_ca+9)}{start}<=100),3,"0")))'
        # aw.cell(row=start, column=start_col_ca+10).value=f'=IF(AND({get_column_letter(start_col_ca+3)}{start}>{start_range_1},{get_column_letter(start_col_ca+3)}{start}<{end_range_1}),{map_1},IF(AND({get_column_letter(start_col_ca+3)}{start}>{start_range_2},{get_column_letter(start_col_ca+3)}{start}<{end_range_2}),{map_2},IF(AND({get_column_letter(start_col_ca+3)}{start}>{start_range_3},{get_column_letter(start_col_ca+3)}{start}<{end_range_3}),{map_3},"0")))'
        target_cell=aw.cell(row=start, column=start_col_ca+10)
        target_cell.value=(
            f'=IF(AND({get_column_letter(start_col_ca+9)}{start}>={start_range_1},'
            f'{get_column_letter(start_col_ca+9)}{start}<={end_range_1}),{map_1},'
            f'IF(AND({get_column_letter(start_col_ca+9)}{start}>{end_range_1},'
            f'{get_column_letter(start_col_ca+9)}{start}<={end_range_2}),{map_2},'
            f'IF(AND({get_column_letter(start_col_ca+9)}{start}>{end_range_2},'
            f'{get_column_letter(start_col_ca+9)}{start}<={end_range_3}),{map_3},"0")))'
        )


        direct_attainment = f'{get_column_letter(start_col_ca+7)}{start}'
        indirect_attainment = f'{get_column_letter(start_col_ca+9)}{start}'
        calculation = f'={direct_attainment}*(B17/100)+{indirect_attainment}*(B18/100)'
        formula = f'={calculation}'
        aw.cell(row=start, column=start_col_ca+11).value = formula

        # aw.cell(row=start, column=start_col_ca+12).value=f'=IF(AND({get_column_letter(start_col_ca+11)}{start}>0,{get_column_letter(start_col_ca+11)}{start}<40),1,IF(AND({get_column_letter(start_col_ca+11)}{start}>=40,{get_column_letter(start_col_ca+11)}{start}<60),2,IF(AND({get_column_letter(start_col_ca+11)}{start}>=60,{get_column_letter(start_col_ca+11)}{start}<=100),3,"0")))'
        # aw.cell(row=start, column=start_col_ca+12).value=f'=IF(AND({get_column_letter(start_col_ca+3)}{start}>{start_range_1},{get_column_letter(start_col_ca+3)}{start}<{end_range_1}),{map_1},IF(AND({get_column_letter(start_col_ca+3)}{start}>{start_range_2},{get_column_letter(start_col_ca+3)}{start}<{end_range_2}),{map_2},IF(AND({get_column_letter(start_col_ca+3)}{start}>{start_range_3},{get_column_letter(start_col_ca+3)}{start}<{end_range_3}),{map_3},"0")))'
        target_cell=aw.cell(row=start, column=start_col_ca+12)
        target_cell.value=(
            f'=IF(AND({get_column_letter(start_col_ca+11)}{start}>={start_range_1},'
            f'{get_column_letter(start_col_ca+11)}{start}<={end_range_1}),{map_1},'
            f'IF(AND({get_column_letter(start_col_ca+11)}{start}>{end_range_1},'
            f'{get_column_letter(start_col_ca+11)}{start}<={end_range_2}),{map_2},'
            f'IF(AND({get_column_letter(start_col_ca+11)}{start}>{end_range_2},'
            f'{get_column_letter(start_col_ca+11)}{start}<={end_range_3}),{map_3},"0")))'
        )


        rowindex+=1
        start=start+interval


    cellstyle_range(aw[f'{get_column_letter(start_col_ca)}{start_row_ca}:{get_column_letter(start_col_ca+12)}{aw.max_row}'],border=True, alignment=True, wrap_text=True)
            

    #================================================================================================================================================================
    current_row = aw.max_row+4
    current_col = start_col_ca
    # aw.merge_cells(start_row=current_row, start_column=current_col, end_row=current_row, end_column=17+current_col)
    aw.merge_cells(start_row=current_row, start_column=current_col, end_row=current_row, end_column=config["PO"]+config["PSO"]+current_col)
    aw.cell(row=current_row, column=current_col).value = "Weighted PO/PSO Attainment Contribution"
    cellstyle_range(aw[f'{get_column_letter(current_col)}{current_row}:{get_column_letter(config["PO"]+config["PSO"]+current_col)}{current_row}'],border=True, fill="ffe74e", bold=True, alignment=True)

    current_row+=1
    aw.cell(row=current_row, column=current_col).value = "COs\\POs"
    cellstyle(aw.cell(row=current_row, column=current_col), bold=True, alignment=True, border=True, fill="4bacc6", font_color="FFFFFF")

    # for popso in range(1,12+5+1):
    for popso in range(1,config["PO"]+config["PSO"]+1):
        if popso<=config["PO"]:
            aw.cell(row=current_row, column=popso+current_col).value = f"PO{popso}"
        else:
            aw.cell(row=current_row, column=popso+current_col).value = f"PSO{popso-config['PO']}"
        cellstyle(aw.cell(row=current_row, column=popso+current_col), bold=True, alignment=True, border=True, fill="4bacc6", font_color="FFFFFF")

    current_row+=1

    start=start_row_ca+3
    # interval=17
    interval=config["PO"]+config["PSO"]

    current_col=start_col_ca
    for co in range(1,data["Number_of_COs"]+1):
        aw[f"{get_column_letter(current_col)}{current_row}"]=f"CO{co}"
        cellstyle(aw[f"{get_column_letter(current_col)}{current_row}"], bold=True, alignment=True, border=True, fill="4bacc6", font_color="FFFFFF")
        

 
        # for popso in range(1,12+5+1):
        for popso in range(1,config["PO"]+config["PSO"]+1):
            aw[f"{get_column_letter(popso+current_col)}{current_row}"]=f'={get_column_letter(start_col_ca+2)}{start+popso}*{get_column_letter(start_col_ca+12)}{start+1}'
            cellstyle(aw[f"{get_column_letter(popso+current_col)}{current_row}"], alignment=True, border=True)

        #current_col=4; popso=1; current_row=80
        #col=6 row=23


        current_row+=1
        start+=interval

    current_col=1
    aw.cell(row=current_row, column=current_col).value = "Academic Year"
    aw.cell(row=current_row+1, column=current_col).value = data["Academic_year"]
    
    current_col+=1
    aw.cell(row=current_row, column=current_col).value = "Semester"
    aw.cell(row=current_row+1, column=current_col).value = data["Semester"]
    
    current_col+=1
    aw.cell(row=current_row, column=current_col).value = "Subject Name"
    aw.cell(row=current_row+1, column=current_col).value = data["Subject_Name"]
    
    current_col+=1
    aw.cell(row=current_row, column=current_col).value = "Subject Code"
    aw.cell(row=current_row+1, column=current_col).value = data["Subject_Code"]

    
    current_col+=1
    # aw.merge_cells(start_row=current_row, start_column=current_col, end_row=current_row, end_column=16+current_col)
    aw.merge_cells(start_row=current_row, start_column=current_col, end_row=current_row, end_column=config["PO"]+config["PSO"]-1+current_col)
    aw.cell(row=current_row, column=current_col).value = "Final Ratio"

    cellstyle_range(aw[f'{get_column_letter(current_col-4)}{current_row}:{get_column_letter(config["PO"]+config["PSO"]-1+current_col)}{current_row}'],border=True, fill="ffe74e", bold=True, alignment=True)
    cellstyle_range(aw[f'{get_column_letter(current_col-4)}{current_row+1}:{get_column_letter(current_col-1)}{current_row+1}'],border=True, fill="4bacc6", bold=True, alignment=True, font_color="FFFFFF")
    current_row+=1
    current_col=4
    

    # for popso in range(1,12+5+1):
    for popso in range(1,config["PO"]+config["PSO"]+1):
        main_formula = f'SUM({get_column_letter(popso+current_col)}{current_row-1-data["Number_of_COs"]}:{get_column_letter(popso+current_col)}{current_row-2})/(SUM({get_column_letter(popso+4)}3:{get_column_letter(popso+4)}{data["Number_of_COs"]+2}))'
        complete_formula = f'=IF(AND(SUM({get_column_letter(popso+current_col)}{current_row-1-data["Number_of_COs"]}:{get_column_letter(popso+current_col)}{current_row-2})>0, SUM({get_column_letter(popso+4)}3:{get_column_letter(popso+4)}{data["Number_of_COs"]+2})>0), {main_formula}, 0)'
        aw[f"{get_column_letter(popso+current_col)}{current_row}"] = complete_formula
        cellstyle(aw[f"{get_column_letter(popso+current_col)}{current_row}"], alignment=True, border=True, bold=True)


    return aw