from openpyxl.styles import Font                                          #import font and alignment from openpyxl
from openpyxl.utils import  get_column_letter
from .utils import adjust_width, cellstyle_range, cellstyle
from .Input_Details import input_detail

def printout(aw, data,config,start_row,copy=False,numco=None):

    #============================================================================================================
    if not copy:
        aw = input_detail(data, {}, aw, copy=True)
        adjust_width(aw)

    if numco is None:
        numco = data['Number_of_COs']
    # ===========================================================================================================     
    headingrow=start_row-1
    aw.merge_cells(f'D{headingrow}:R{headingrow}')
    if len(data) > 0:
        aw[f'D{headingrow}'] = f"{data['Section']}_{data['Batch']}_{data['Branch']}_{data['Semester']}_{data['Subject_Code']}"
    cellstyle(aw[f'D{headingrow}'], bold=True, alignment=True, border=True, fill="ce875c",size=14)
    #============================================================================================================

    start_column=4

    column=start_column
    row=start_row
    # Merging cells dynamically based on row and column number
    aw.merge_cells(start_row=row, end_row=row+2, start_column=column,  end_column=column)
    # Setting value, font, and alignment for the merged cell
    cell_reference = f"{get_column_letter(column)}{row}"
    aw[cell_reference] = "Course Code"
    aw.merge_cells(start_row=row+3, end_row=row+3+numco-1, start_column=column,  end_column=column)
    
    aw.column_dimensions[f"{get_column_letter(column)}"].width = 8.43
    column+=1
    #============================================================================================================
    # Merging cells for "Course Name"
    aw.merge_cells(start_row=row, end_row=row+2, start_column=column,  end_column=column)

    # Setting value, font, and alignment for "Course Name" cell
    cell_reference = f"{get_column_letter(column)}{row}"
    aw[cell_reference] = "Course Name"
     
    aw.merge_cells(start_row=row+3, end_row=row+3+numco-1, start_column=column,  end_column=column)

    aw.column_dimensions[f"{get_column_letter(column)}"].width = 8.43
    column+=1

    #============================================================================================================            
    interval=config["PO"]+config["PSO"]
    
    # Merging cells for "COs"
    aw.merge_cells(start_row=row, end_row=row+2, start_column=column,  end_column=column)
    cell_reference = f"{get_column_letter(column)}{row}"
    aw[cell_reference] = "COs"
    for nco in range(numco):
        aw[f"{get_column_letter(column)}{row+3+nco}"] = f"CO{nco+1}"
        if nco%2==0:
            cellstyle(aw[f"{get_column_letter(column)}{row+3+nco}"], fill="ffff00")

    aw.column_dimensions[f"{get_column_letter(column)}"].width = 12
    column+=1          
    #============================================================================================================
    aw.merge_cells(start_row=row, end_row=row, start_column=column, end_column=column+1)
    ese_cell_reference = f"{get_column_letter(column)}{row}"
    aw[ese_cell_reference] = "End Semester Examination"
    
    #============================================================================================================
    aw.merge_cells(start_row=row+1, end_row=row+1, start_column=column, end_column=column+1)
    see_cell_reference = f"{get_column_letter(column)}{row+1}"
    aw[see_cell_reference] = "(SEE)*"


    #============================================================================================================
    
    attainment_cell_reference = f"{get_column_letter(column)}{row+2}"
    aw[attainment_cell_reference] = "Attainment"
    aw.column_dimensions[f"{get_column_letter(column)}"].width = 12
    column+=1
    #============================================================================================================
    level_cell_reference = f"{get_column_letter(column)}{row+2}"
    aw[level_cell_reference] = "Level"
    for nco in range(numco):
        cellstyle(aw[f"{get_column_letter(column)}{row+3+nco}"], fill="fde9d9")

    
    aw.column_dimensions[f"{get_column_letter(column)}"].width = 12
    column+=1
    #============================================================================================================
    aw.merge_cells(start_row=row, end_row=row, start_column=column, end_column=column+1)
    ie_cell_reference = f"{get_column_letter(column)}{row}"
    aw[ie_cell_reference] = "Internal Examination"
    
    #============================================================================================================
    aw.merge_cells(start_row=row+1, end_row=row+1, start_column=column, end_column=column+1)
    cie_cell_reference = f"{get_column_letter(column)}{row+1}"
    aw[cie_cell_reference] = "(CIE)*"
    
    #============================================================================================================        
    
    attainment_cell_reference = f"{get_column_letter(column)}{row+2}"
    aw[attainment_cell_reference] = "Attainment"
    aw.column_dimensions[f"{get_column_letter(column)}"].width = 12
    column+=1

    #============================================================================================================
   
    level_cell_reference = f"{get_column_letter(column)}{row+2}"
    aw[level_cell_reference] = "Level"  
    for nco in range(numco):
        cellstyle(aw[f"{get_column_letter(column)}{row+3+nco}"], fill="fde9d9")
    aw.column_dimensions[f"{get_column_letter(column)}"].width = 12
    column+=1

    #============================================================================================================
    aw.merge_cells(start_row=row, end_row=row, start_column=column, end_column=column+1)

    # Setting value, font, and alignment for "Direct" cell
    direct_cell_reference = f"{get_column_letter(column)}{row}"
    aw[direct_cell_reference] = "Direct"
   
    #============================================================================================================
    aw.merge_cells(start_row=row+1, end_row=row+1, start_column=column, end_column=column+1)

    # Setting value, font, and alignment for "CIE + SEE" cell
    cie_see_cell_reference = f"{get_column_letter(column)}{row+1}"
    aw[cie_see_cell_reference] = f"=B15 & \" % of CIE + \" & B16 & \" % of SEE\""

    #============================================================================================================
    attainment_cell_reference = f"{get_column_letter(column)}{row+2}"
    aw[attainment_cell_reference] = "Attainment"  
    aw.column_dimensions[f"{get_column_letter(column)}"].width = 12
    column+=1
        
    #============================================================================================================
    level_cell_reference = f"{get_column_letter(column)}{row+2}"
    aw[level_cell_reference] = "Level"
    for nco in range(numco):
        cellstyle(aw[f"{get_column_letter(column)}{row+3+nco}"], fill="fde9d9")
    aw.column_dimensions[f"{get_column_letter(column)}"].width = 12
    column+=1
    #============================================================================================================
    aw.merge_cells(start_row=row, end_row=row+1, start_column=column, end_column=column+1)
    indirect_cell_reference = f"{get_column_letter(column)}{row}"
    aw[indirect_cell_reference] = "Indirect"

    #============================================================================================================
    attainment_cell_reference = f"{get_column_letter(column)}{row+2}"
    aw[attainment_cell_reference] = "Attainment"
    aw.column_dimensions[f"{get_column_letter(column)}"].width = 12    
    column+=1
    #============================================================================================================
    level_cell_reference = f"{get_column_letter(column)}{row+2}"
    aw[level_cell_reference] = "Level"
    for nco in range(numco):
        cellstyle(aw[f"{get_column_letter(column)}{row+3+nco}"], fill="fde9d9")
    aw.column_dimensions[f"{get_column_letter(column)}"].width = 8.43
    column+=1
    #============================================================================================================
    aw.merge_cells(start_row=row, end_row=row, start_column=column, end_column=column+1)
    total_course_attainment_cell_reference = f"{get_column_letter(column)}{row}"
    aw[total_course_attainment_cell_reference] = "Total Course Attainment"

    #============================================================================================================
    aw.merge_cells(start_row=row+1, end_row=row+1, start_column=column, end_column=column+1)
    direct_indirect_cell_reference = f"{get_column_letter(column)}{row+1}"
    aw[direct_indirect_cell_reference] = f"=B17 & \" % of Direct + \" & B18 & \" % of Indirect\""

    #============================================================================================================
    attainment_cell_reference = f"{get_column_letter(column)}{row+2}"
    aw[attainment_cell_reference] = "Attainment"
    aw.column_dimensions[f"{get_column_letter(column)}"].width = 20                            
    column+=1
    #============================================================================================================
    level_cell_reference = f"{get_column_letter(column)}{row+2}"
    aw[level_cell_reference] = "Level"
    for nco in range(numco):
        cellstyle(aw[f"{get_column_letter(column)}{row+3+nco}"], fill="fde9d9")
    aw.column_dimensions[f"{get_column_letter(column)}"].width = 8.43
    column+=1
    #============================================================================================================
    target_cell_reference = f"{get_column_letter(column)}{row}"
    aw[target_cell_reference] = "Target"
    
    #============================================================================================================
    percentage_cell_reference = f"{get_column_letter(column)}{row+1}"
    aw[percentage_cell_reference] = "(%)"
    aw.column_dimensions[f"{get_column_letter(column)}"].width = 8.43
    for nco in range(numco):
        cellstyle(aw[f"{get_column_letter(column)}{row+3+nco}"], fill="ffff00")
    column+=1
    #============================================================================================================
    final_attainment_cell_reference = f"{get_column_letter(column)}{row}"
    aw[final_attainment_cell_reference] = "Final Attainment"

    #============================================================================================================
    yesno_cell_reference = f"{get_column_letter(column)}{row+1}"
    aw[yesno_cell_reference] = "Yes/No"
    aw[yesno_cell_reference].font = Font(bold=True)
    aw.column_dimensions[f"{get_column_letter(column)}"].width = 20
    column+=1
    #============================================================================================================

    #Printing all the data in the excel sheet
    if copy == False:
        start_row_ca_data=numco+8+numco+3+4
        start_col_ca_data=4+3
        start_column=4
        row=start_row+3
        aw[f"{get_column_letter(start_column)}{row}"] = data['Subject_Code']
        aw[f"{get_column_letter(start_column+1)}{row}"] = data['Subject_Name']
        for nco in range(numco):
            aw[f"{get_column_letter(start_column+3)}{row}"] = f"={data['Section']}_Course_Attainment!{get_column_letter(start_col_ca_data)}{start_row_ca_data+(nco*interval)}"
            aw[f"{get_column_letter(start_column+4)}{row}"] = f"={data['Section']}_Course_Attainment!{get_column_letter(start_col_ca_data+1)}{start_row_ca_data+(nco*interval)}"
            aw[f"{get_column_letter(start_column+5)}{row}"] = f"={data['Section']}_Course_Attainment!{get_column_letter(start_col_ca_data+2)}{start_row_ca_data+(nco*interval)}"
            aw[f"{get_column_letter(start_column+6)}{row}"] = f"={data['Section']}_Course_Attainment!{get_column_letter(start_col_ca_data+3)}{start_row_ca_data+(nco*interval)}"
            aw[f"{get_column_letter(start_column+7)}{row}"] = f"={data['Section']}_Course_Attainment!{get_column_letter(start_col_ca_data+4)}{start_row_ca_data+(nco*interval)}"
            aw[f"{get_column_letter(start_column+8)}{row}"] = f"={data['Section']}_Course_Attainment!{get_column_letter(start_col_ca_data+5)}{start_row_ca_data+(nco*interval)}"
            aw[f"{get_column_letter(start_column+9)}{row}"] = f"={data['Section']}_Course_Attainment!{get_column_letter(start_col_ca_data+6)}{start_row_ca_data+(nco*interval)}"
            aw[f"{get_column_letter(start_column+10)}{row}"] = f"={data['Section']}_Course_Attainment!{get_column_letter(start_col_ca_data+7)}{start_row_ca_data+(nco*interval)}"
            aw[f"{get_column_letter(start_column+11)}{row}"] = f"={data['Section']}_Course_Attainment!{get_column_letter(start_col_ca_data+8)}{start_row_ca_data+(nco*interval)}"
            aw[f"{get_column_letter(start_column+12)}{row}"] = f"={data['Section']}_Course_Attainment!{get_column_letter(start_col_ca_data+9)}{start_row_ca_data+(nco*interval)}"
            aw[f"{get_column_letter(start_column+13)}{row}"] = f"=B19"
            aw[f"{get_column_letter(start_column+14)}{row}"] = f'=IF({get_column_letter(start_column+11)}{row}>=B19,"Yes","No")'
            row+=1


   
    #============================================================================================================

    cellstyle_range(aw[f"D{start_row}:R{3+start_row+numco-1}"],border=True, bold=True, alignment=True, wrap_text=True)
    cellstyle_range(aw[f"G{start_row+2}:R{start_row+2}"],border=True, bold=True, alignment=True, wrap_text=True, fill="8db4e2")
    cellstyle(aw[f"D{start_row+3}"],border=True, bold=True, alignment=True, wrap_text=True, text_rotation=90, fill="ffff00")
    cellstyle(aw[f"E{start_row+3}"],border=True, bold=True, alignment=True, wrap_text=True, text_rotation=90, fill="1ed760")
    return aw