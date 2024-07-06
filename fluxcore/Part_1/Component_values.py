from openpyxl.styles import PatternFill                                           #import patternfill from openpyxl
from openpyxl.utils import get_column_letter                                  #import get_column_letter from openpyxl
from openpyxl.formatting.rule import FormulaRule
from .utils import adjust_width, colour_table, cellstyle_range, cellstyle
from openpyxl.styles import Protection

def qn_co_mm_btl(data,key,value,aw):  #function to create qn_co_mm_btl table
    """ Function to create qn_co_mm_btl table

    Args:
    data (dict): Dictionary containing the data
    key (str): Component name
    value (int): Number of questions
    aw (openpyxl.worksheet.worksheet.Worksheet): Worksheet object

    Returns:
    openpyxl.worksheet.worksheet.Worksheet: Worksheet object
    """
    for row in aw.iter_rows():
            for cell in row:
                cell.protection = Protection(locked=True)
    aw.protection.sheet = True
        
    aw.merge_cells(f'B1:{get_column_letter(value+2)}1')
    aw[f'B1']=key
    cellstyle_range(aw[f'B1:{get_column_letter(value+2)}1'], bold=True, alignment=True, border=True, fill="ffe74e")
    
    aw['B2']="Question"
    aw['B3']="Max Marks"
    aw['B4']="Threshold"
    aw['B5']="CO"
    aw['B6']="Final CO"
    aw['B7']="BTL"

    cellstyle_range(aw[f'B2:B7'], bold=True, alignment=True, border=True, fill="4bacc6")
    

    for qno in range(1,value+1):
        aw[f'{get_column_letter(qno+2)}2']=f"Q{qno}"
        cellstyle(aw[get_column_letter(qno+2)+'2'], bold=True, alignment=True, border=True, fill="4bacc6")
        
        aw[f'{get_column_letter(qno+2)}4']=f'={data["Section"]}_Input_Details!B14/100*{get_column_letter(qno+2)}3'
        
        aw[f'{get_column_letter(qno+2)}6'].value = f'=CONCATENATE("{data["Subject_Code"]+"_CO"}", {get_column_letter(qno+2)}5)'
        cellstyle(aw[get_column_letter(qno+2)+'6'], bold=True)


    cellstyle_range(aw[f'C3:{get_column_letter(value+2)}7'], border=True, alignment=True, alternate=['daeef3', 'b7dee8'])

    pink_fill = PatternFill(start_color="D8A5B5", end_color="D8A5B5", fill_type="solid")
    light_red_fill = PatternFill(start_color="FF5E5E", end_color="FF5E5E", fill_type="solid")
    for qno in range(1, value + 1):
        max_marks_cell = f'{get_column_letter(qno+2)}3'
        threshold_cell = f'{get_column_letter(qno+2)}4'
        co_cell = f'{get_column_letter(qno+2)}5'
        btl_cell = f'{get_column_letter(qno+2)}7'

        #set conditional formatting for max marks cell such that if its more than 100, it will be highlighted red
        aw.conditional_formatting.add(
            max_marks_cell,
            FormulaRule(formula=[f'OR({max_marks_cell}>100,{max_marks_cell}<0)'], stopIfTrue=False, fill=light_red_fill))
        aw.conditional_formatting.add(
            max_marks_cell,
            FormulaRule(formula=[f'ISBLANK({max_marks_cell})'], stopIfTrue=False, fill=pink_fill))
        
        # Conditional Formatting Rule
        aw.conditional_formatting.add(
            threshold_cell,
            FormulaRule(formula=[f'OR({threshold_cell}>max_marks_cell,{threshold_cell}<0)'], stopIfTrue=False, fill=light_red_fill))
        aw.conditional_formatting.add(
            threshold_cell,
            FormulaRule(formula=[f'ISBLANK({threshold_cell})'], stopIfTrue=False, fill=pink_fill))

        # Conditional Formatting Rule for co cell such that if its more than data["Number_of_COs"], it will be highlighted red
        aw.conditional_formatting.add(
            co_cell,
            FormulaRule(formula=[f'OR({co_cell}>{data["Number_of_COs"]},{co_cell}<0)'], stopIfTrue=False, fill=light_red_fill))
        aw.conditional_formatting.add(
            co_cell,
            FormulaRule(formula=[f'ISBLANK({co_cell})'], stopIfTrue=False, fill=pink_fill))

        # Conditional Formatting Rule for btl cell such that if its more than 100, it will be highlighted red
        aw.conditional_formatting.add(
            btl_cell,
            FormulaRule(formula=[f'OR({btl_cell}>100,{btl_cell}<0)'], stopIfTrue=False, fill=light_red_fill))
        aw.conditional_formatting.add(
            btl_cell,
            FormulaRule(formula=[f'ISBLANK({btl_cell})'], stopIfTrue=False, fill=pink_fill))
        
    #unprotecting the cells
    purple_fill = PatternFill(start_color="D8A5B5", end_color="D8A5B5", fill_type="solid")
    for row in aw.iter_rows(min_row=3, max_row=7, min_col=3, max_col=value+2):
        for cell in row:
            if cell.row != 6:
                cell.protection = Protection(locked=False)
                #cell.fill = purple_fill

    return aw

def studentmarks(data,key, value,aw):
    """ Function to create studentmarks table

    Args:
    data (dict): Dictionary containing the data
    key (str): Component name
    value (int): Number of questions
    aw (openpyxl.worksheet.worksheet.Worksheet): Worksheet object

    Returns:
    openpyxl.worksheet.worksheet.Worksheet: Worksheet object
    """

    aw.merge_cells(f'B9:{get_column_letter(value+2)}9')
    aw["B9"]="Marks obtained"
    cellstyle_range(aw[f'A9:{get_column_letter(value+2)}9'], bold=True, alignment=True, border=True, fill="ffe74e")

    aw["A10"]="Roll No."
    aw["B10"]="Name"

    for qno in range(1,value+1):
        aw[get_column_letter(qno+2)+'10']=f"Q{qno}"
        
    cellstyle_range(aw[f'A10:{get_column_letter(value+2)}10'], bold=True, alignment=True, border=True, fill="4bacc6")
    cellstyle_range(aw[f'A11:{get_column_letter(value+2)}{10+data["Number_of_Students"]}'], border=True, alignment=True, alternate=['daeef3', 'b7dee8'])


    #conditional formatting
    pink_fill = PatternFill(start_color="D8A5B5", end_color="D8A5B5", fill_type="solid")
    light_red_fill = PatternFill(start_color="FF5E5E", end_color="FF5E5E", fill_type="solid")
    yellow_fill = PatternFill(start_color="d9a46f", end_color="d9a46f", fill_type="solid")
    # Apply conditional formatting to each question's column header
    for qno in range(1, value + 1):
        column_letter = get_column_letter(qno + 2)
        header_cell = f"${column_letter}$10"
        data_range = f"{column_letter}11:{column_letter}{10 + data['Number_of_Students']}"
        threshold_cell = f"${column_letter}$4"
        max_marks_cell = f"${column_letter}$3"  # Assuming max marks is in row 3

        # The formula checks if all cells in the range below the header are below the threshold
        
        # Since the COUNTIF will return count of all cells not less than the threshold, we only apply
        # the format if the result of COUNTIF is 0, meaning all filled cells are below the threshold
        aw.conditional_formatting.add(
            header_cell,
            FormulaRule(formula=[f"COUNTIF({data_range}, \">=\"&{threshold_cell})=0"], stopIfTrue=False, fill=yellow_fill)
        )
       
        #Highlight empty cells
        aw.conditional_formatting.add(
            data_range,
            FormulaRule(formula=[f'ISBLANK({column_letter}11)'], stopIfTrue=False, fill=pink_fill)
        )

        #Highlight cells with value greater than max marks
        aw.conditional_formatting.add(
            data_range,
            FormulaRule(formula=[f'{column_letter}11>{max_marks_cell}'], stopIfTrue=False, fill=light_red_fill)
        )

        #highlight roll no. and name cells with pink if empty
        aw.conditional_formatting.add(
            f"A11:A{10 + data['Number_of_Students']}",
            FormulaRule(formula=[f'ISBLANK(A11)'], stopIfTrue=False, fill=pink_fill)
        )
        aw.conditional_formatting.add(
            f"B11:B{10 + data['Number_of_Students']}",
            FormulaRule(formula=[f'ISBLANK(B11)'], stopIfTrue=False, fill=pink_fill)
        )

    #unprotecting the cells
    purple_fill = PatternFill(start_color="D8A5B5", end_color="D8A5B5", fill_type="solid")
    for row in aw.iter_rows(min_row=11, max_row=10+data['Number_of_Students'], min_col=1, max_col=value+2):
        for cell in row:
            cell.protection = Protection(locked=False)
            #cell.fill = purple_fill
   
    adjust_width(aw)
    aw.column_dimensions['A'].width = 20
    aw.column_dimensions['B'].width = 30

    colour_table(aw, data)

    return aw