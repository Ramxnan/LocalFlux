from openpyxl.utils import get_column_letter                                  #import get_column_letter from openpyxl
from openpyxl.styles import Border, Side, Alignment
from .utils import cellstyle, cellstyle_range


def cummulative_qn_co_mm_btl(data, component_name, numquestions, aw):
    """ Function to create cummulative_co_mm_btl table

    Args:
    data (dict): Dictionary containing the data
    component_name (str): Component name
    numquestions (int): Number of questions
    aw (openpyxl.worksheet.worksheet.Worksheet): Worksheet object

    Returns:
    openpyxl.worksheet.worksheet.Worksheet: Worksheet object
    """


    for cno in range(1, data["Number_of_COs"]+1):
        aw[f'{get_column_letter(numquestions+3+cno)}2'] = f'CO{cno}'
        cellstyle(aw[f'{get_column_letter(numquestions+3+cno)}2'], bold=True, fill="4bacc6", font_color="FFFFFF", border=True, alignment=True)
       
        co_name = f'{data["Subject_Code"]}_CO{cno}'  # Replace with the actual CO name you're checking
        sum_range_start_maxmarks = 'C3'
        sum_range_end_maxmarks = f'{get_column_letter(numquestions + 2)}3'
        criteria_range_start = 'C6'
        criteria_range_end = f'{get_column_letter(numquestions + 2)}6'

        aw[f'{get_column_letter(numquestions+3+cno)}3'] = f'=SUMIFS({sum_range_start_maxmarks}:{sum_range_end_maxmarks}, {criteria_range_start}:{criteria_range_end}, "{co_name}")'

        sum_range_start_threshold = 'C4'
        sum_range_end_threshold = f'{get_column_letter(numquestions + 2)}4'
        aw[f'{get_column_letter(numquestions+3+cno)}4'] = f'=SUMIFS({sum_range_start_threshold}:{sum_range_end_threshold}, {criteria_range_start}:{criteria_range_end}, "{co_name}")'
        
        cellstyle_range(aw[f'{get_column_letter(numquestions+3+cno)}3:{get_column_letter(numquestions+3+cno)}4'], alignment=True, border=True)
    
    return aw

def cummulative_studentmarks(data, component_name, numquestions, aw):
    """ Function to create cummulative_studentmarks table

    Args:
    data (dict): Dictionary containing the data
    component_name (str): Component name
    numquestions (int): Number of questions
    aw (openpyxl.worksheet.worksheet.Worksheet): Worksheet object

    Returns:
    openpyxl.worksheet.worksheet.Worksheet: Worksheet object
    """
    for cno in range(1, data["Number_of_COs"]+1):
        aw[f'{get_column_letter(numquestions+3+cno)}10'] = f'CO{cno}'
        cellstyle(aw[f'{get_column_letter(numquestions+3+cno)}10'], bold=True, fill="4bacc6", font_color="FFFFFF", border=True, alignment=True)

        co_name = f'{data["Subject_Code"]}_CO{cno}'
        criteria_range_start = 'C6'
        criteria_range_end = f'{get_column_letter(numquestions + 2)}6'

        for nstudents in range(1, data["Number_of_Students"]+1):
            sum_range_start_marks = f'C{10+nstudents}'
            sum_range_end_marks = f'{get_column_letter(numquestions + 2)}{10+nstudents}'
            aw[f'{get_column_letter(numquestions+3+cno)}{10+nstudents}'] = f'=SUMIFS({sum_range_start_marks}:{sum_range_end_marks}, {criteria_range_start}:{criteria_range_end}, "{co_name}")'
            cellstyle(aw[f'{get_column_letter(numquestions+3+cno)}{10+nstudents}'], alignment=True, border=True)


    return aw
