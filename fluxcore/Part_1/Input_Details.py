
from .utils import colour_table_Input_Details, cellstyle, cellstyle_range
from openpyxl.styles import PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo          #import table from openpyxl
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter                                  #import get_column_letter from openpyxl
from openpyxl.styles import Protection


def input_detail(data,Component_Details,aw,id_page=False, copy=False):
    for row in aw.iter_rows():
        for cell in row:
            cell.protection = Protection(locked=True)
    aw.protection.sheet = True
    """ Function to input details

    Args:
    data (dict): Dictionary containing the data
    Component_Details (dict): Dictionary containing the component details
    aw (openpyxl.worksheet.worksheet.Worksheet): Worksheet object

    Returns:
    openpyxl.worksheet.worksheet.Worksheet: Worksheet object
    """

    aw.merge_cells('A1:B1')
    aw['A1']="Constants"
    cellstyle_range(aw['A1:B1'], bold=True, alignment=True, border=True, fill="ffe74e")

    startrow=2
    for key, value in data.items():
        aw[f'A{startrow}']=key
        aw[f'B{startrow}']=value
        startrow+=1

    cellstyle_range(aw[f'A2:B{startrow-1}'], border=True, alignment=True, bold=True, alternate=['b7dee8', 'daeef3'])
    startrow+=1

    aw.merge_cells(f'A{startrow}:B{startrow}')
    aw[f'A{startrow}']="Variables"
    cellstyle_range(aw[f'A{startrow}:B{startrow}'], bold=True, alignment=True, border=True, fill="ffe74e")
    
    internal_components_number = len([key for key in Component_Details.keys() if key.endswith("I")])
    external_components_number = len([key for key in Component_Details.keys() if key.endswith("E")])

    aw['A14']="Default Threshold %"
    aw['A15']="Internal %"
    aw['A16']="External %"
    if internal_components_number==0:
        aw['B15']=f'0'
    elif external_components_number==0:
        aw['B15']=f'100'
    aw['B16']=f'=100-B15'
    aw['A17']="Direct %"
    aw['A18']="Indirect %"
    aw['B18']=f'=100-B17'
    aw['A19']="Target CO Attainment %"

    cellstyle_range(aw[f'A14:B19'], border=True, alignment=True, bold=True, alternate=['b7dee8', 'daeef3'])

    if copy:
        #copy the values from the previous sheet
        aw['B14']=f"='{data['Section']}_Input_Details'!B14"
        aw['B15']=f"='{data['Section']}_Input_Details'!B15"
        aw['B16']=f"='{data['Section']}_Input_Details'!B16"
        aw['B17']=f"='{data['Section']}_Input_Details'!B17"
        aw['B18']=f"='{data['Section']}_Input_Details'!B18"
        aw['B19']=f"='{data['Section']}_Input_Details'!B19"


    # =================================================================================================================================================================
    if id_page:
        aw['A22']="Component Details"
        aw['B22']="Number of Questions"
        cellstyle_range(aw['A22:B22'], bold=True, alignment=True, border=True, font_color="FFFFFF")

        startrow=23
        for key, value in Component_Details.items():
            aw[f'A{startrow}']=key
            aw[f'B{startrow}']=value
            cellstyle_range(aw[f'A{startrow}:B{startrow}'], alignment=True, border=True, bold=True)
            startrow+=1
        
        #make a table
        tab = Table(displayName=f"{data['Section']}_Component_Details", ref=f"A22:B{aw.max_row}")
        style = TableStyleInfo(name="TableStyleMedium15", showFirstColumn=False,
                            showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        aw.add_table(tab)
    
        pink_fill = PatternFill(start_color="D8A5B5", end_color="D8A5B5", fill_type="solid")
        red_fill = PatternFill(start_color="ff5e5e", end_color="ff5e5e", fill_type="solid")
        #set conditional formatting for B9 to B19 such that if its empty, it will be highlighted pink
        for startrow in range(14,20):
            if startrow!=16 and startrow!=18:
                aw.conditional_formatting.add(
                    f'B{startrow}',
                    FormulaRule(formula=[f'ISBLANK(B{startrow})'], stopIfTrue=False, fill=pink_fill))
                aw.conditional_formatting.add(
                    f'B{startrow}',
                    #greater than 100 or less than 0
                    FormulaRule(formula=[f'OR(B{startrow}>100,B{startrow}<0)'], stopIfTrue=False, fill=red_fill))
        
        colour_table_Input_Details(aw)

            

        #unlock the cells
        purple_fill = PatternFill(start_color="7b83eb", end_color="7b83eb", fill_type="solid")
        for row in aw.iter_rows(min_row=14, max_row=19, min_col=2, max_col=2):
            for cell in row:
                if cell.row != 16 and cell.row != 18:
                    cell.protection = Protection(locked=False)
                    #cell.fill = purple_fill

    return aw  



def CO_PO_Table(data,config,aw,id_page=False, copy=False):
    """
    Function to create CO-PO Table

    Args:
    data (dict): Dictionary containing the data
    aw (openpyxl.worksheet.worksheet.Worksheet): Worksheet object

    Returns:
    openpyxl.worksheet.worksheet.Worksheet: Worksheet object
    """
    #merge cells depending on number of POs
    aw.merge_cells(start_row=1, start_column=4, end_row=1, end_column=config["PO"]+config["PSO"]+1+3)
    aw['D1']="CO-PO Mapping"
    cellstyle(aw['D1'], bold=True, alignment=True, border=True, fill="ffe74e")
   
    aw["D2"]="COs\\POs"
    cellstyle(aw["D2"], bold=True, alignment=True, border=True, fill="9bbb59")


    for nco in range(1,data["Number_of_COs"]+1):
        aw[f"D{nco+2}"]=f"CO{nco}"
        cellstyle(aw[f"D{nco+2}"], bold=True, alignment=True, border=True)

    for popso in range(1,config["PO"]+config["PSO"]+1):
    # for popso in range(1,12+5+1):
        if popso<=config["PO"]:
            aw[f"{get_column_letter(popso+4)}2"]=f"PO{popso}   "
        else:
            aw[f"{get_column_letter(popso+4)}2"]=f"PSO{popso-config['PO']}   "
        cellstyle(aw[f"{get_column_letter(popso+4)}2"], bold=True, alignment=True, border=True, fill="9bbb59")
        #set column width to 13
        aw.column_dimensions[f"{get_column_letter(popso+4)}"].width = 13

    cellstyle_range(aw[f"D2:{get_column_letter(config['PO']+config['PSO']+1+3)}2"], bold=True, alignment=True, border=True, fill="9bbb59")


    aw[f'D{3+data["Number_of_COs"]}']=data["Subject_Code"]
    for popso in range(1,config["PO"]+config["PSO"]+1):
        range_string = f"{get_column_letter(popso+4)}3:{get_column_letter(popso+4)}{2+data['Number_of_COs']}"
        aw[f"{get_column_letter(popso+4)}{3+data['Number_of_COs']}"]=f'=IF(SUM({range_string})=0,"",AVERAGE({range_string}))'

    cellstyle_range(aw[f"D{3+data['Number_of_COs']}:{get_column_letter(config['PO']+config['PSO']+1+3)}{3+data['Number_of_COs']}"],border=True, alignment=True, bold=True, fill="92d050")
    cellstyle_range(aw[f"D3:{get_column_letter(config['PO']+config['PSO']+1+3)}{2+data['Number_of_COs']}"],border=True, alternate=['ebf1de','ffffff'], alignment=True)

    
    if id_page:
        #set conditional formatting for empty cells   
        pink_fill = PatternFill(start_color="D8A5B5", end_color="D8A5B5", fill_type="solid")
        red_fill = PatternFill(start_color="ff5e5e", end_color="ff5e5e", fill_type="solid")
        for nco in range(1,data["Number_of_COs"]+1):
            for popso in range(1,config["PO"]+config["PSO"]+1):
                aw.conditional_formatting.add(
                    f"{get_column_letter(popso+4)}{nco+2}",
                    FormulaRule(formula=[f'ISBLANK({get_column_letter(popso+4)}{nco+2})'], stopIfTrue=False, fill=pink_fill))
                aw.conditional_formatting.add(
                    f"{get_column_letter(popso+4)}{nco+2}",
                    FormulaRule(formula=[f'OR({get_column_letter(popso+4)}{nco+2}>3,{get_column_letter(popso+4)}{nco+2}<0)'], stopIfTrue=False, fill=red_fill))
                
        #unlock the cells
        purple_fill = PatternFill(start_color="7b83eb", end_color="7b83eb", fill_type="solid")
        for row in aw.iter_rows(min_row=3, max_row=2+data["Number_of_COs"], min_col=5, max_col=config["PO"]+config["PSO"]+1+3):
            for cell in row:
                cell.protection = Protection(locked=False)
                #cell.fill = purple_fill      
    if copy:
        for nco in range(1,data["Number_of_COs"]+1):
            # for popso in range(1,12+5+1):
            for popso in range(1,config["PO"]+config["PSO"]+1):
                aw[f"{get_column_letter(popso+4)}{nco+2}"]=f"='{data['Section']}_Input_Details'!{get_column_letter(popso+4)}{nco+2}"
                cellstyle(aw[f"{get_column_letter(popso+4)}{nco+2}"], alignment=True, bold=True)
            
    return aw



def indirect_co_assessment(data,aw,id_page=False,copy=False, ca_page=False):
    """
    Function to create Indirect CO Assessment

    Args:
    data (dict): Dictionary containing the data
    aw (openpyxl.worksheet.worksheet.Worksheet): Worksheet object

    Returns:
    openpyxl.worksheet.worksheet.Worksheet: Worksheet object
    """
    #merge cells depending on number of POs
    aw.merge_cells(start_row=data["Number_of_COs"]+5, start_column=4, end_row=data["Number_of_COs"]+5, end_column=5)
    aw[f'D{data["Number_of_COs"]+5}']="Indirect CO Assessment"
    cellstyle_range(aw[f"D{data['Number_of_COs']+5}:E{data['Number_of_COs']+5}"], bold=True, alignment=True, border=True, fill="ffe74e")
    

    aw[f"D{data['Number_of_COs']+6}"]="COs"    
    aw[f"E{data['Number_of_COs']+6}"]="Indirect %"
    cellstyle_range(aw[f"D{data['Number_of_COs']+6}:E{data['Number_of_COs']+6}"], bold=True, alignment=True, border=True, fill="f79646")
   
    
    for nco in range(1,data["Number_of_COs"]+1):
        aw[f"D{nco+data['Number_of_COs']+6}"]=f"CO{nco}"
        cellstyle(aw[f"D{nco+data['Number_of_COs']+6}"], bold=True, alignment=True, border=True)
    
    startrow = data["Number_of_COs"]+7
    endrow = startrow + data["Number_of_COs"]-1
    cellstyle_range(aw[f"D{startrow}:E{endrow}"], alignment=True, border=True, alternate=['fcd5b4', 'fde9d9'])

    if id_page:
        pink_fill = PatternFill(start_color="D8A5B5", end_color="D8A5B5", fill_type="solid")
        red_fill = PatternFill(start_color="ff5e5e", end_color="ff5e5e", fill_type="solid")
        for row in range(startrow, endrow+1):
            cell_coordinate = f'E{row}'
            aw.conditional_formatting.add(
                cell_coordinate,
                FormulaRule(formula=[f'ISBLANK({cell_coordinate})'], stopIfTrue=False, fill=pink_fill))
            aw.conditional_formatting.add(
                cell_coordinate,
                #greater than 100 or less than 0
                FormulaRule(formula=[f'OR({cell_coordinate}>100,{cell_coordinate}<0)'], stopIfTrue=False, fill=red_fill))
            
        #unlock the cells
        purple_fill = PatternFill(start_color="7b83eb", end_color="7b83eb", fill_type="solid")
        for row in aw.iter_rows(min_row=data["Number_of_COs"]+7, max_row=data["Number_of_COs"]+6+data["Number_of_COs"], min_col=5, max_col=5):
            for cell in row:
                cell.protection = Protection(locked=False)
                #cell.fill = purple_fil
    if copy:
        for nco in range(1,data["Number_of_COs"]+1):
            aw[f"E{nco+data['Number_of_COs']+6}"]=f"='{data['Section']}_Input_Details'!E{nco+data['Number_of_COs']+6}"
            cellstyle(aw[f"E{nco+data['Number_of_COs']+6}"], alignment=True, bold=True)


    if ca_page:
        startrow = data["Number_of_COs"]+5


        # Merge cells and add headers
        aw.merge_cells(f'G{startrow}:I{startrow}')  
        aw[f'G{startrow}'] = "Attainment Mapping"
        cellstyle_range(aw[f'G{startrow}:I{startrow}'], bold=True, alignment=True, border=True, fill="ffe74e")

        startrow += 1
        aw[f'G{startrow}'] = "Start Range"
        aw[f'H{startrow}'] = "End Range"
        aw[f'I{startrow}'] = "Mapping"
        cellstyle_range(aw[f'G{startrow}:I{startrow}'], bold=True, alignment=True, border=True, fill="8a6f9f")

        # Add range and mapping data
        startrow += 1
        aw[f'G{startrow}'] = 0
        aw[f'H{startrow}'] = 40
        aw[f'I{startrow}'] = 1

        startrow += 1
        aw[f'G{startrow}'] = 40.1
        aw[f'H{startrow}'] = 60
        aw[f'I{startrow}'] = 2

        startrow += 1
        aw[f'G{startrow}'] = 60.1
        aw[f'H{startrow}'] = 100
        aw[f'I{startrow}'] = 3

        cellstyle_range(aw[f'G{startrow-2}:I{startrow}'], alignment=True, border=True, bold=True, alternate=['ffffff', 'e4dfec'])

        #add conditional formatting
        pink_fill = PatternFill(start_color="d8a5b5", end_color="d8a5b5", fill_type="solid")
        red_fill = PatternFill(start_color="ff5e5e", end_color="ff5e5e", fill_type="solid")
        
        for row in range(startrow-2, startrow+1):
            for col in range(7, 9):
                # If cell is empty or less than 0 or greater than 100 then highlight it with pink or red
                aw.conditional_formatting.add(
                    f"{get_column_letter(col)}{row}",
                    FormulaRule(formula=[f'OR({get_column_letter(col)}{row}>100, {get_column_letter(col)}{row}<0)'], stopIfTrue=False, fill=red_fill)
                )
                aw.conditional_formatting.add(
                    f"{get_column_letter(col)}{row}",
                    FormulaRule(formula=[f'ISBLANK({get_column_letter(col)}{row})'], stopIfTrue=False, fill=pink_fill)
                )
                #if the row is start-2 and the cell in col 7 does not start with 0, highlight it with red
                if row == startrow-2 and col == 7:
                    aw.conditional_formatting.add(
                        f"{get_column_letter(col)}{row}",
                        FormulaRule(formula=[f'{get_column_letter(col)}{row}<>0'], stopIfTrue=False, fill=red_fill)
                    )
                #if the row is startrow and the cell in col 8 is not 100, highlight it with red
                if row == startrow and col == 8:
                    aw.conditional_formatting.add(
                        f"{get_column_letter(col)}{row}",
                        FormulaRule(formula=[f'{get_column_letter(col)}{row}<>100'], stopIfTrue=False, fill=red_fill)
                    )

        #unlock the cells
        purple_fill = PatternFill(start_color="7b83eb", end_color="7b83eb", fill_type="solid")
        for row in aw.iter_rows(min_row=startrow-2, max_row=startrow, min_col=7, max_col=9):
            for cell in row:
                cell.protection = Protection(locked=False)
                #cell.fill = purple_fill

        
               

                

    return aw