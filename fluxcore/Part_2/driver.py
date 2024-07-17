import sys
sys.path.append('fluxcore')

import sys
import os

# path = os.path.expanduser('~/NBA/nba')
# if path not in sys.path:
#     sys.path.append(path)

from openpyxl import Workbook
from openpyxl.styles import Protection
from Part_1.utils import adjust_width, cellstyle_range
from Part_1.Input_Details import input_detail, indirect_co_assessment, CO_PO_Table
from Part_1.Component_values import qn_co_mm_btl, studentmarks
from Part_1.Cummulative_Component_Values import cummulative_qn_co_mm_btl, cummulative_studentmarks
from Part_1.InternalExternal_Component_calculation import Component_calculation
from Part_1.write_course_attainment import write_course_attainment
from Part_1.printout import printout
import os
import uuid

import os
import uuid
from openpyxl import load_workbook
import pandas as pd
from openpyxl.utils import get_column_letter

def driver_part2(input_dir_path, output_dir_path, config):


    #create openpyxl workbook
    wbwrite = Workbook()
    wbwrite.remove(wbwrite.active)
    excel_files=[]
    total_students = 0
    Warnings = []


    Combined_co_po_table = pd.DataFrame()
    Combined_qn_co_mm_btl = {}
    Combined_studentmarks = {}

    prev_qn_co_mm_btl_check = {}
    prev_co_po_table_check = None
    prev_Component_Details_check = None

    new_qn_co_mm_btl_check = {}
    new_co_po_table_check = None
    new_Component_Details_check = None

    Input_Details_Sheet_Array = []

    for file in os.listdir(input_dir_path):
        if file.endswith(".xlsx") and not file.startswith("Combined"):
            excel_files.append(file)
    excel_files.sort()

    #All the files should have different sections
    sections = []
    for file in excel_files:
        sections.append(file[0])

    #if section is not letter, return an error
    for section in sections:
        if not section.isalpha():
            Warnings.append(f"{file} has invalid section name")
            return Warnings

    if len(sections) != len(set(sections)):
        Warnings.append("All the files should have different sections")
        return Warnings


    for file in excel_files:
        if file.endswith(".xlsx") and not file.startswith("Combined"):
            file_path = os.path.join(input_dir_path, file)
            wbread = load_workbook(file_path, data_only=True)

            alldata={}
            Component_Details = {}


            input_details_title=None
            for sheet_name in wbread.sheetnames:
                if sheet_name.endswith("Input_Details"):
                    input_details_title = sheet_name
                    Input_Details_Sheet_Array.append(sheet_name)

            if input_details_title is None:
                Warnings.append(f"{file} does not have Input_Details sheet")
                return Warnings

            wsread_input_details = wbread[input_details_title]
            #create a dictionary from A2 to B11
            for key, value in wsread_input_details.iter_rows(min_row=2, max_row=11, min_col=1, max_col=2, values_only=True):
                alldata[key] = value
    
            for key, value in wsread_input_details.iter_rows(min_row=14, max_row=19, min_col=1, max_col=2, values_only=True):
                alldata[key] = value

            #if any of the values in alldata is None, return an error
            if None in alldata.values():
                Warnings.append(f"{file} has missing values in Constants or Variables in Input_Details sheet")
                return Warnings

            total_students+=alldata['Number_of_Students']
            data = {key: alldata[key] for key in alldata.keys() & {'Teacher', 'Academic_year', 'Batch', 'Branch', 'Subject_Name', 'Subject_Code', 'Section', 'Semester', 'Number_of_Students', 'Number_of_COs'}}
            Combined_data_all = alldata

            
                
            #extract table called Component_Details and store it in a dictionary
            table_range = wsread_input_details.tables[f'{data["Section"]}_Component_Details'].ref
            for row in wsread_input_details[table_range][1:]:
                Component_Details[row[0].value] = row[1].value

            Combined_Component_Details = {f"Combined_{key[2:]}":value for key,value in Component_Details.items()}

            #Check if Component_Details is same as previous file
            new_Component_Details_check = {f"{key[2:]}":value for key,value in Component_Details.items()}
            if prev_Component_Details_check is not None:
                for key in new_Component_Details_check.keys():
                    if prev_Component_Details_check[key] != new_Component_Details_check[key]:
                        Warnings.append(f"{file} has different Component_Details")
                        return Warnings
            prev_Component_Details_check = new_Component_Details_check

            #Check if the number of PO and PSO is same as config
            start_copo = 5
            end_copo = wsread_input_details.max_column
            if end_copo-start_copo+1 != config["PO"]+config["PSO"]:
                Warnings.append(f"{file} has different number of PO and PSO as set in config")
                return Warnings


            new_qn_co_mm_btl_check = {}         
            for key, numques in Component_Details.items():
                comp_ws = wbread[key]

                #=================QN-CO-MM-BTL Table=================
                # Define the range for QN-CO-MM-BTL details
                start_row = 2
                end_row = 7
                start_col = 3
                end_col = 3 + numques - 1
                
                # Extract QN-CO-MM-BTL details
                data_rows = []
                for row in comp_ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
                    data_rows.append([cell.value for cell in row])
                df_qn_co_mm_btl = pd.DataFrame(data_rows[1:], columns=data_rows[0])
                Combined_qn_co_mm_btl["Combined_"+key[2:]] = df_qn_co_mm_btl
                
                new_qn_co_mm_btl_check[key[2:]] = df_qn_co_mm_btl


                # Define the range for student marks
                start_row = 10
                end_row = 10 + alldata['Number_of_Students']
                start_col = 1
                end_col = 2 + numques

                # Extract student marks
                data_rows = []
                for row in comp_ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
                    data_rows.append([cell.value for cell in row])
                df_student_marks = pd.DataFrame(data_rows[1:], columns=data_rows[0])
                if "Combined_"+key[2:] in Combined_studentmarks:
                    Combined_studentmarks["Combined_"+key[2:]] = pd.concat([Combined_studentmarks["Combined_"+key[2:]], df_student_marks], axis=0)
                else:
                    Combined_studentmarks["Combined_"+key[2:]] = df_student_marks
                 
            #Check if QN-CO-MM-BTL is same as previous file
            if prev_qn_co_mm_btl_check:
                for key in new_qn_co_mm_btl_check.keys():
                    if not prev_qn_co_mm_btl_check[key].equals(new_qn_co_mm_btl_check[key]):
                        Warnings.append(f"{file} has different QN-CO-MM-BTL Table in {key} Component")
                        return Warnings
            prev_qn_co_mm_btl_check = new_qn_co_mm_btl_check

            
            #=================CO-PO Table=================
            start_row = 3
            end_row = 3 + data["Number_of_COs"] - 1
            start_col = 5
            end_col=start_col+config["PO"]+config["PSO"]-1
            cell_range = f'{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}'

            values = []
            for row in wsread_input_details[cell_range]:
                values.append([cell.value for cell in row])

            # po_columns=[f"PO{i}" for i in range(1,13)]
            # pso_columns=[f"PSO{i}" for i in range(1,6)]
            po_columns=[f"PO{i}" for i in range(1,config["PO"]+1)]
            pso_columns=[f"PSO{i}" for i in range(1,config["PSO"]+1)]
            df_columns=po_columns+pso_columns
            
            co_po_table=pd.DataFrame(values, columns=df_columns)
            #replace 0 with NaN
            co_po_table.replace(0, pd.NA, inplace=True)
            Combined_co_po_table = co_po_table

            new_co_po_table_check = co_po_table
            if prev_co_po_table_check is not None:
                if not prev_co_po_table_check.equals(new_co_po_table_check):
                    Warnings.append(f"{file} has different CO-PO Table")
                    return Warnings
            prev_co_po_table_check = new_co_po_table_check
            

            #=================Write to Combined Workbook=================
            wbwrite.create_sheet(f"{data['Section']}_Input_Details")
            wswrite = wbwrite[f"{data['Section']}_Input_Details"]
            wswrite = input_detail(data,Component_Details,wswrite, id_page=True)
            wswrite = indirect_co_assessment(data,wswrite, id_page=True)
            adjust_width(wswrite)
            wswrite = CO_PO_Table(data,config,wswrite, id_page=True)

            for key in Component_Details.keys():
                wbwrite.create_sheet(key)
                wswrite = wbwrite[key]
                wswrite.title = key
                wswrite = qn_co_mm_btl(data, key, Component_Details[key], wswrite)
                wswrite = studentmarks(data, key, Component_Details[key], wswrite)

                wswrite = cummulative_qn_co_mm_btl(data, key, Component_Details[key], wswrite)   
                wswrite = cummulative_studentmarks(data, key, Component_Details[key], wswrite)

            internal_components_number = len([key for key in Component_Details.keys() if key.endswith("I")])
            external_components_number = len([key for key in Component_Details.keys() if key.endswith("E")])

            
            wbwrite.create_sheet(f"{data['Section']}_Internal_Components")
            wswrite = wbwrite[f"{data['Section']}_Internal_Components"]
            if internal_components_number==0:
                wswrite.merge_cells('A1:M1')
                cellstyle_range(wswrite['A1:M1'], bold=True, alignment=True, border=True, fill="ffe74e", size=18)
                wswrite['A1']="No Internal Components"

                wswrite.merge_cells('A2:M2')
                cellstyle_range(wswrite['A2:M2'], bold=True, alignment=True, border=True, fill="ffe74e", size=12)
                wswrite['A2']="Ensure that Internal % is set to 0 since there are no Internal Components"
            else:
                wswrite = Component_calculation(data,Component_Details,wswrite,"I")

            wbwrite.create_sheet(f"{data['Section']}_External_Components")
            wswrite = wbwrite[f"{data['Section']}_External_Components"]
            if external_components_number==0:
                wswrite.merge_cells('A1:M1')
                cellstyle_range(wswrite['A1:M1'], bold=True, alignment=True, border=True, fill="ffe74e", size=18)
                wswrite['A1']="No External Components"

                wswrite.merge_cells('A2:M2')
                cellstyle_range(wswrite['A2:M2'], bold=True, alignment=True, border=True, fill="ffe74e", size=12)
                wswrite['A2']="Ensure that External % is set to 0 since there are no External Components"
            else:
                wswrite = Component_calculation(data,Component_Details,wswrite,"E")

            wbwrite.create_sheet(f"{data['Section']}_Course_Attainment")
            wswrite = wbwrite[f"{data['Section']}_Course_Attainment"]
            wswrite=write_course_attainment(data, Component_Details, config, wswrite)

            wbwrite.create_sheet(f"{data['Section']}_Printout")
            wswrite = wbwrite[f"{data['Section']}_Printout"]
            wswrite=printout(wswrite,data,config,2)

            #copy data from all the sheets of wbread to wbwrite
            for sheet in wbread.sheetnames:
                wsread = wbread[sheet]
                wswrite = wbwrite[sheet]
                for row in wsread.iter_rows(min_row=1, max_row=wsread.max_row, min_col=1, max_col=wsread.max_column):
                    for cell in row:
                        #if error occurs while copying the cell, skip the cell
                        try:
                            wswrite[cell.coordinate].value = cell.value
                        except:
                            pass


            wbread.close()
    #save the workbook
    
    Combined_data_all['Section'] = "Combined"
    Combined_data_all['Number_of_Students'] = total_students
    Combined_data = {key: Combined_data_all[key] for key in Combined_data_all.keys() & {'Teacher', 'Academic_year', 'Semester', 'Branch', 'Batch', 'Section', 'Subject_Code', 'Subject_Name', 'Number_of_Students', 'Number_of_COs','Default Threshold %','Internal %','Direct %','Target CO Attainment %'}}
    #order it in the same order as the input details
    Combined_data = {key: Combined_data_all[key] for key in ['Teacher', 'Academic_year', 'Semester','Branch','Batch','Section','Subject_Code' ,'Subject_Name', 'Number_of_Students', 'Number_of_COs']}

    # print(total_students)
    # print(Combined_data_all)
    # print(Combined_Component_Details)
    # print(Combined_co_po_table)
    # print(Combined_qn_co_mm_btl)
    # print(Combined_studentmarks)

    wbwrite.create_sheet(f"{Combined_data['Section']}_Input_Details")
    wswrite = wbwrite[f"{Combined_data['Section']}_Input_Details"]
    wswrite['B14'] = Combined_data_all['Default Threshold %']
    wswrite['B15'] = Combined_data_all['Internal %']
    wswrite['B17'] = Combined_data_all['Direct %']
    wswrite['B19'] = Combined_data_all['Target CO Attainment %']

    
    #calculate average of indirect assessment
    for numco in range(Combined_data['Number_of_COs']):
        formula="=AVERAGE("
        for ws in Input_Details_Sheet_Array:
            formula+=f"{ws}!E{2+Combined_data['Number_of_COs']+4+1+numco},"
        formula = formula[:-1]
        formula+=")"
        wswrite[f'E{2+Combined_data["Number_of_COs"]+4+1+numco}'] = formula
        
    wswrite = input_detail(Combined_data,Combined_Component_Details,wswrite)
    
    

    wswrite = indirect_co_assessment(Combined_data,wswrite)
    adjust_width(wswrite)

    start_row = 3
    end_row = 3 + data["Number_of_COs"] - 1
    start_col = 5
    end_col=start_col+config["PO"]+config["PSO"]-1
    #paste the content of Combined_co_po_table to given range in the sheet
    row=0
    col=0
    for r in range(start_row, end_row+1):
        col=0
        for c in range(start_col, end_col+1):
            wswrite.cell(row=r, column=c, value=Combined_co_po_table.iloc[row,col])
            col+=1
        row+=1

    wswrite = CO_PO_Table(Combined_data,config,wswrite)




    for key in Combined_Component_Details.keys():
        #replace first letter with combined
        wbwrite.create_sheet(key)
        wswrite = wbwrite[key]
        wswrite.title = key
        wswrite = qn_co_mm_btl(Combined_data, key, Combined_Component_Details[key], wswrite)
        #paste the content of Combined_qn_co_mm_btl to given range in the sheet
        start_row = 3
        end_row = 7
        start_col = 3
        end_col = 3 + Combined_Component_Details[key] - 1
        row=0
        col=0
        try:
            for r in range(start_row, end_row+1):
                col=0
                for c in range(start_col, end_col+1):
                    wswrite.cell(row=r, column=c, value=Combined_qn_co_mm_btl[key].iloc[row,col])
                    col+=1
                row+=1
        except:
            pass

        wswrite = studentmarks(Combined_data, key, Combined_Component_Details[key], wswrite)
        start_row = 11
        end_row = 10 + Combined_data['Number_of_Students']
        start_col = 1
        end_col = 2 + Combined_Component_Details[key]
        row=0
        col=0
        try:
            for r in range(start_row, end_row+1):
                col=0
                for c in range(start_col, end_col+1):
                    wswrite.cell(row=r, column=c, value=Combined_studentmarks[key].iloc[row,col])
                    col+=1
                row+=1
        except:
            pass

        wswrite = cummulative_qn_co_mm_btl(Combined_data, key, Combined_Component_Details[key], wswrite)
        wswrite = cummulative_studentmarks(Combined_data, key, Combined_Component_Details[key], wswrite)

  
    

    internal_components_number = len([key for key in Combined_Component_Details.keys() if key.endswith("I")])
    external_components_number = len([key for key in Combined_Component_Details.keys() if key.endswith("E")])

    
    wbwrite.create_sheet(f"{Combined_data['Section']}_Internal_Components")
    wswrite = wbwrite[f"{Combined_data['Section']}_Internal_Components"]
    if internal_components_number==0:
        wswrite.merge_cells('A1:M1')
        cellstyle_range(wswrite['A1:M1'], bold=True, alignment=True, border=True, fill="ffe74e", size=18)
        wswrite['A1']="No Internal Components"

        wswrite.merge_cells('A2:M2')
        cellstyle_range(wswrite['A2:M2'], bold=True, alignment=True, border=True, fill="ffe74e", size=12)
        wswrite['A2']="Ensure that Internal % is set to 0 since there are no Internal Components"
    else:
        wswrite = Component_calculation(Combined_data,Combined_Component_Details,wswrite,"I")

    wbwrite.create_sheet(f"{Combined_data['Section']}_External_Components")
    wswrite = wbwrite[f"{Combined_data['Section']}_External_Components"]
    if external_components_number==0:
        wswrite.merge_cells('A1:M1')
        cellstyle_range(wswrite['A1:M1'], bold=True, alignment=True, border=True, fill="ffe74e", size=18)
        wswrite['A1']="No External Components"

        wswrite.merge_cells('A2:M2')
        cellstyle_range(wswrite['A2:M2'], bold=True, alignment=True, border=True, fill="ffe74e", size=12)
        wswrite['A2']="Ensure that External % is set to 0 since there are no External Components"
    else:
        wswrite = Component_calculation(Combined_data,Combined_Component_Details,wswrite,"E")

    wbwrite.create_sheet("Combined_Course_Attainment")
    wswrite = wbwrite["Combined_Course_Attainment"]
    wswrite=write_course_attainment(Combined_data, Combined_Component_Details,config, wswrite)

    wbwrite.create_sheet("Combined_Printout")
    wswrite = wbwrite["Combined_Printout"]
    wswrite=printout(wswrite,Combined_data,config,2)

    unique_id = str(uuid.uuid4()).split("-")[0]
    excel_file_name=f"Combined_{data['Batch']}_{data['Branch']}_{data['Semester']}_{data['Subject_Code']}_{unique_id}.xlsx"
    
    #Remove all Protection from the workbook
    for sheet in wbwrite.sheetnames:
        wswrite = wbwrite[sheet]
        if sheet.endswith("Printout"):
            wswrite.protection.sheet = False

    wbwrite.save(os.path.join(output_dir_path, excel_file_name))
    if Warnings:
        return Warnings
    else:
        return ["Files successfully merged under File name: " + excel_file_name]
        #return excel_file_name
    
if __name__ == "__main__":
    input_dir_path="C:\\Users\\raman\\OneDrive - Amrita vishwa vidyapeetham\\ASE\\Projects\\NBA\\NBA_v3\\dev_19.1\\flux\\nba\\Part_2"
    output_dir_path="C:\\Users\\raman\\OneDrive - Amrita vishwa vidyapeetham\\ASE\\Projects\\NBA\\NBA_v3\\dev_19.1\\flux\\nba\\Part_2"
    driver_part2(input_dir_path, output_dir_path)