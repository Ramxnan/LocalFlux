from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from openpyxl import Workbook                                                         #import workbook from openpyxl
from openpyxl.styles import  PatternFill, Font, Protection
import pandas as pd
import os
import numpy as np
import uuid
from openpyxl.utils import get_column_letter
from Part_1.printout import printout
from Part_1.utils import cellstyle_range, cellstyle
#from printout import printout_template


def driver_part3(input_dir_path, output_dir_path,config):
    warnings = []
    #check if there are any files in the input directory
    if not os.listdir(input_dir_path):
        warnings.append("No files found in the input directory")
        return warnings
    



    
    wbwrite = Workbook()
    wbwrite.remove(wbwrite.active)
    wbwrite.create_sheet("Printouts",0)
    wbwrite.create_sheet("PO_calculations",1)
    
    wswrite_printouts=wbwrite["Printouts"]
    wswrite_printouts.merge_cells('D1:R1')
    wswrite_printouts["D1"]="Summary of Printouts"
    cellstyle_range(wswrite_printouts['D1:R1'], size=18, bold=True, alignment=True, fill='ffe74e', border=True)

    startrow=3
    for file in os.listdir(input_dir_path):
        if file.endswith(".xlsx") and not file.startswith("final"):
            wbread = load_workbook(os.path.join(input_dir_path, file), data_only=True)
            ws_printout=""
            for ws in wbread.sheetnames:
                if ws.endswith("Printout") and ws.startswith("Combined"):
                    ws_printout=ws

            if ws_printout=="":
                warnings.append(f"Printout sheet not found in {file}")
                return warnings
        

            




            wsread_printout=wbread[ws_printout]
            Number_of_COs=wsread_printout["B11"].value
            #print('Worksheet name:', ws_printout)
            #print('File name:', file)

            wswrite_printouts=printout(wswrite_printouts,{},config,startrow,numco=Number_of_COs,copy=True)
            

            min_row=1
            min_col=4
            max_row=1+3+Number_of_COs
            max_col=18
            for row in range(min_row, max_row+1):
                for col in range(min_col, max_col+1):
                    try:
                        wswrite_printouts.cell(row=startrow-1, column=col).value=wsread_printout.cell(row=row, column=col).value
                        #print("Row:", startrow-1, "Column:", col, "Value:", wsread_printout.cell(row=row, column=col).value)
                    except:
                        #print(f"Error in {file} at row {row} and column {col}")
                        pass
                startrow+=1
          



            startrow+=1
            

            
    #================================================================================================
    #================================================================================================
    # #PO calculation
    wswrite_POCalculation=wbwrite["PO_calculations"]

    wswrite_POCalculation.merge_cells(f'A1:{get_column_letter(3+config['PO']+config['PSO'])}1')
    wswrite_POCalculation["A1"]="PO Attainment"
    cellstyle_range(wswrite_POCalculation[f'A1:{get_column_letter(3+config['PO']+config['PSO'])}1'], size=18, bold=True, alignment=True, fill='ffe74e', border=True)

    wswrite_POCalculation.merge_cells(f'A2:{get_column_letter(3+config['PO']+config['PSO'])}2')
    wswrite_POCalculation["A2"]="Direct Attainment at PO level"
    cellstyle_range(wswrite_POCalculation[f'A2:{get_column_letter(3+config['PO']+config['PSO'])}2'], size=14, bold=True, alignment=True, border=True)
 
    wswrite_POCalculation["A3"]="S.No"   
    wswrite_POCalculation["B3"]="Course Code"
    wswrite_POCalculation["C3"]="Course Name"

    po_data_row=3
    po_data_col=3
    # for popso in range(1,12+5+1):
    for popso in range(1,config['PO']+config['PSO']+1):
        if popso<=config['PO']:
            wswrite_POCalculation.cell(row=po_data_row, column=popso+po_data_col).value=f"PO{popso}"
        else:
            wswrite_POCalculation.cell(row=po_data_row, column=popso+po_data_col).value=f"PSO{popso-config['PO']}"

    cellstyle_range(wswrite_POCalculation[f'A{po_data_row}:{get_column_letter(3+config['PO']+config['PSO'])}{po_data_row}'], bold=True, alignment=True, border=True, fill='6cb266', font_color='ffffff')

    columns=[]
    columns.append("Academic Year")
    columns.append("Semester")
    columns.append("Course Code")
    columns.append("Course Name")
    # for popso in range(1,12+5+1):
    for popso in range(1,config['PO']+config['PSO']+1):
        if popso<=config['PO']:
            columns.append(f"PO{popso}")
        else:
            columns.append(f"PSO{popso-config['PO']}")
    final_po_table=pd.DataFrame(columns=columns)

    for file in os.listdir(input_dir_path):
        #file shouldnt start with final
        if file.endswith(".xlsx") and not file.startswith("final"):
            wbread = load_workbook(os.path.join(input_dir_path, file), data_only=True)
            #find the name of worksheet which ends with Input_Details
            wsname_ID=""
            wsname_CA=""
            for ws in wbread.sheetnames:
                if ws.endswith("Input_Details"):
                    wsname_ID=ws
                if ws.endswith("Course_Attainment"):
                    wsname_CA=ws

            if wsname_ID=="":
                warnings.append(f"Input_Details sheet not found in {file}")
                return warnings
            if wsname_CA=="":
                warnings.append(f"Course_Attainment sheet not found in {file}")
                return warnings
            
            # Check if number of POs and PSOs are same as in config file
            start_copo = 5
            end_copo = wbread[wsname_ID].max_column
            if end_copo-start_copo+1 != config["PO"]+config["PSO"]:
                warnings.append(f"{file} has different number of PO and PSO as set in config")
                return warnings

            wsread_input_detials=wbread[wsname_ID]
            Number_of_COs=wsread_input_detials["B11"].value

            wsread_Course_Attainment=wbread[wsname_CA]
            row=wsread_Course_Attainment.max_row
            min_col=1
            max_col=4+config["PO"]+config["PSO"]
            rowdata=[]
            for col in range(min_col, max_col+1):
                    rowdata.append(wsread_Course_Attainment.cell(row=row, column=col).value)
            rowdata_df=pd.DataFrame(rowdata).T
            rowdata_df.columns=columns
            final_po_table = pd.concat([final_po_table,rowdata_df], axis=0)
            
            
    final_po_table=final_po_table.replace(0, np.nan)
    final_po_table.reset_index(drop=True, inplace=True)

    semester_sort = {'Odd': 1, 'Even': 2}
    final_po_table['Semester code'] = final_po_table['Semester'].map(semester_sort)

    final_po_table = final_po_table.sort_values(by=['Academic Year', 'Semester code'])
    #print(final_po_table)

    dataframes_dict = {group: data.drop(['Academic Year', 'Semester', 'Semester code'], axis=1)
                   for group, data in final_po_table.groupby(['Academic Year', 'Semester'])}



    for key, value in dataframes_dict.items():
        print(key)
        print(value)
        print("=====================================")

    startrow=4
    startcol=1
    sno=1
    trows=[]
    vrows=[]
    for key, value in dataframes_dict.items():
        wswrite_POCalculation.merge_cells(start_row=startrow, start_column=startcol, end_row=startrow, end_column=startcol+config['PO']+config['PSO']+2)
        wswrite_POCalculation.cell(row=startrow, column=startcol).value=f"{key[0]} {key[1]}"
        cellstyle_range(wswrite_POCalculation[f'A{startrow}:{get_column_letter(3+config['PO']+config['PSO'])}{startrow}'], bold=True, alignment=True, fill='b7dee8', border=True)

        startrow+=1
        ridex=0
        for _ in dataframe_to_rows(value, index=False, header=False):
            vrows.append(startrow)
            cindex = 0
            for c in range(2, startcol+3+config['PO']+config['PSO']):
                wswrite_POCalculation.cell(row=startrow, column=1).value=f'{sno}.'
                cellstyle(wswrite_POCalculation.cell(row=startrow, column=1), alignment=True, border=True)

                wswrite_POCalculation.cell(row=startrow, column=c).value=value.iloc[ridex, cindex]
                cellstyle(wswrite_POCalculation.cell(row=startrow, column=c), alignment=True, border=True)

                if ridex==len(value)-1 and c == startcol+3+config['PO']+config['PSO']-1:
                    startrow+=1
                    trows.append(startrow)
                    wswrite_POCalculation.merge_cells(start_row=startrow, start_column=startcol, end_row=startrow, end_column=startcol+2)
                    wswrite_POCalculation.cell(row=startrow, column=startcol).value="Total"
                    for cfin in range (4, 4+config['PO']+config['PSO']):
                        wswrite_POCalculation.cell(row=startrow, column=cfin).value=f"=SUM({get_column_letter(cfin)}{startrow-ridex-1}:{get_column_letter(cfin)}{startrow-1})"
                    cellstyle_range(wswrite_POCalculation[f'A{startrow}:{get_column_letter(3+config['PO']+config['PSO'])}{startrow}'], alignment=True, border=True, fill='fcd5b4', bold=True)
                    startrow+=1

                cindex+=1

            startrow+=1
            sno+=1
            ridex+=1

    # #================================================================================================
    wswrite_POCalculation.merge_cells(f'A{startrow}:{get_column_letter(3+config['PO']+config['PSO'])}{startrow}')
    wswrite_POCalculation[f"A{startrow}"]="Indirect Assessment At PO Level"
    cellstyle_range(wswrite_POCalculation[f'A{startrow}:{get_column_letter(3+config['PO']+config['PSO'])}{startrow}'], size=14, bold=True, alignment=True, border=True)
  
    po_data_col=3
    startrow+=1

    # for popso in range(1,12+5+1):
    for popso in range(1,config['PO']+config['PSO']+1):
        if popso<=config['PO']:
            wswrite_POCalculation.cell(row=startrow, column=popso+po_data_col).value=f"PO{popso}"
        else:
            wswrite_POCalculation.cell(row=startrow, column=popso+po_data_col).value=f"PSO{popso-config['PO']}"
    cellstyle_range(wswrite_POCalculation[f'A{startrow}:{get_column_letter(3+config['PO']+config['PSO'])}{startrow}'], bold=True, alignment=True, fill='6CB266', border=True, font_color='ffffff')
    
    startrow+=1
    wswrite_POCalculation[f'A{startrow}']=f'{sno}.'
    sno+=1
    wswrite_POCalculation.merge_cells(start_row=startrow, end_row=startrow, start_column=2, end_column=3)
    wswrite_POCalculation[f'B{startrow}']="Exit survey feedback"
    
    startrow+=1
    wswrite_POCalculation[f'A{startrow}']=f'{sno}.'
    sno+=1
    wswrite_POCalculation.merge_cells(start_row=startrow, end_row=startrow, start_column=2, end_column=3)
    wswrite_POCalculation[f'B{startrow}']="Recruiters Feedback"
    
    cellstyle_range(wswrite_POCalculation[f'A{startrow-1}:{get_column_letter(3+config['PO']+config['PSO'])}{startrow}'], alignment=True, border=True)

    startrow+=1
    wswrite_POCalculation.merge_cells(start_row=startrow, end_row=startrow, start_column=1, end_column=3)
    wswrite_POCalculation[f'A{startrow}']="Average"
    for colind in range(4,4+config['PO']+config['PSO']):
        wswrite_POCalculation[f'{get_column_letter(colind)}{startrow}']=f'=IFERROR(AVERAGE({get_column_letter(colind)}{startrow-2}:{get_column_letter(colind)}{startrow-1}),0)'

    cellstyle_range(wswrite_POCalculation[f'A{startrow}:{get_column_letter(3+config['PO']+config['PSO'])}{startrow}'], alignment=True, border=True, fill='fcd5b4', bold=True)


    # #================================================================================================
        
    startrow+=2
    wswrite_POCalculation.merge_cells(f'A{startrow}:{get_column_letter(3+config['PO']+config['PSO'])}{startrow}')
    wswrite_POCalculation[f"A{startrow}"]='Total PO Attainment'
    cellstyle_range(wswrite_POCalculation[f'A{startrow}:{get_column_letter(3+config['PO']+config['PSO'])}{startrow}'], size=18, bold=True, alignment=True, fill='95b3d7', border=True)

    startrow+=1
    wswrite_POCalculation.cell(row=startrow, column=1).fill = PatternFill(start_color='6CB266', end_color='6CB266', fill_type='solid')
    wswrite_POCalculation.cell(row=startrow, column=2).fill = PatternFill(start_color='6CB266', end_color='6CB266', fill_type='solid')
    wswrite_POCalculation.cell(row=startrow, column=3).fill = PatternFill(start_color='6CB266', end_color='6CB266', fill_type='solid')

    # for popso in range(1,12+5+1):
    for popso in range(1,config['PO']+config['PSO']+1):
        if popso<=config['PO']:
            wswrite_POCalculation.cell(row=startrow, column=popso+po_data_col).value=f"PO{popso}"
        else:
            wswrite_POCalculation.cell(row=startrow, column=popso+po_data_col).value=f"PSO{popso-config['PO']}"
    cellstyle_range(wswrite_POCalculation[f'A{startrow}:{get_column_letter(3+config['PO']+config['PSO'])}{startrow}'], bold=True, alignment=True, border=True, font_color='ffffff', fill='6CB266')

    startrow+=1
    wswrite_POCalculation.merge_cells(f'A{startrow}:C{startrow}')
    wswrite_POCalculation[f'A{startrow}']="Total Direct Assessment"  
    for colind in range(4,4+config['PO']+config['PSO']):
        formula=f'=SUM('
        for trow in trows:
            formula+=f'{get_column_letter(colind)}{trow},'
        formula=formula[:-1]
        formula+=')'
        wswrite_POCalculation[f'{get_column_letter(colind)}{startrow}'].value=formula

    startrow+=1
    wswrite_POCalculation.merge_cells(f'A{startrow}:C{startrow}')
    wswrite_POCalculation[f'A{startrow}']="Total courses through PO mapped"
    for colind in range(4,4+config['PO']+config['PSO']):
        formula=f'=COUNT('
        for vrow in vrows:
            formula+=f'{get_column_letter(colind)}{vrow},'
        formula=formula[:-1]
        formula+=')'
        wswrite_POCalculation[f'{get_column_letter(colind)}{startrow}'].value=formula

    startrow+=1
    wswrite_POCalculation.merge_cells(f'A{startrow}:C{startrow}')
    wswrite_POCalculation[f'A{startrow}']="Average of direct Assessment"
    for colind in range(4,4+config['PO']+config['PSO']):
        wswrite_POCalculation[f'{get_column_letter(colind)}{startrow}'].value=f'=IFERROR({get_column_letter(colind)}{startrow-2}/{get_column_letter(colind)}{startrow-1},0)'

    startrow+=1
    wswrite_POCalculation.merge_cells(f'A{startrow}:C{startrow}')
    wswrite_POCalculation[f'A{startrow}']="Average of Indirect Assessment"
    for colind in range(4,4+config['PO']+config['PSO']):
        wswrite_POCalculation[f'{get_column_letter(colind)}{startrow}'].value=f'={get_column_letter(colind)}{startrow-7}'

    cellstyle_range(wswrite_POCalculation[f'A{startrow-3}:{get_column_letter(3+config['PO']+config['PSO'])}{startrow}'], alignment=True, border=True)

    startrow+=1
    wswrite_POCalculation.merge_cells(f'A{startrow}:C{startrow}')
    wswrite_POCalculation[f'A{startrow}']="PO Attainment for the Program"
    wswrite_POCalculation[f'A{startrow}'].font = Font(bold=True, size=14)
    
    cellstyle_range(wswrite_POCalculation[f'A{startrow}:{get_column_letter(3+config['PO']+config['PSO'])}{startrow}'], alignment=True, border=True, bold=True)

    #set column width for second column to be 12
    wswrite_POCalculation.column_dimensions['B'].width = 12
    wswrite_POCalculation.column_dimensions['C'].width = 16

    unique_code = str(uuid.uuid4()).split("-")[0]
    file_name = f"final_{unique_code}.xlsx"
    for ws in wbwrite.sheetnames:
        ws = wbwrite[ws]
        ws.protection.sheet = True
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.protection = Protection(locked=True)

    wbwrite.save(os.path.join(output_dir_path, file_name))
    #return file_name
    if warnings:
        return warnings
    else:
        return ['Batch processing completed successfully under file name: ' + file_name]

if __name__ == "__main__":
    input_dir_path="C:\\Users\\raman\\OneDrive - Amrita vishwa vidyapeetham\\ASE\\Projects\\NBA\\NBA_v3\\dev_19.1\\flux\\nba\\Part_2"
    output_dir_path="C:\\Users\\raman\\OneDrive - Amrita vishwa vidyapeetham\\ASE\\Projects\\NBA\\NBA_v3\\dev_19.1\\flux\\nba\\Part_2"
    driver_part3(input_dir_path, output_dir_path)